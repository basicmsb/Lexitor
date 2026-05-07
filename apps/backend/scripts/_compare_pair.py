"""Print original.xlsx vs ARH.xlsx rows side-by-side for one base pair.

Shows the first 'meaningful' sheet (not naslovnica/sadrzaj/opci uvjeti)
so we can compare how Marko maps raw rows into the canonical Arhigon
format, row by row. Read-only.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from openpyxl import load_workbook  # noqa: E402

ROOT = Path(r"C:\Dropbox\Arhigon WEB\_TVRTKE - Unos podataka\_R_")

SKIP_SHEET_TOKENS = ("naslovnic", "sadrz", "opci uvj", "opće uvj", "rekap")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Side-by-side compare original vs ARH")
    p.add_argument("--orig", type=Path, required=True)
    p.add_argument("--arh", type=Path, required=True)
    p.add_argument(
        "--rows",
        type=int,
        default=40,
        help="How many rows to print from each side (default 40)",
    )
    return p.parse_args()


def first_meaningful_sheet(wb) -> str:
    for name in wb.sheetnames:
        norm = name.lower().strip()
        if any(tok in norm for tok in SKIP_SHEET_TOKENS):
            continue
        ws = wb[name]
        if ws.max_row > 5:
            return name
    return wb.sheetnames[0]


def cell_repr(cell) -> str:
    if cell.value is None:
        return ""
    if cell.data_type == "f":
        return f"=f({cell.value})"
    s = str(cell.value).replace("\n", " ⏎ ")
    return s


def extract_rows(path: Path, n_rows: int) -> tuple[str, list[list[str]]]:
    wb = load_workbook(filename=str(path), data_only=False)
    try:
        sheet_name = first_meaningful_sheet(wb)
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=n_rows):
            rows.append([cell_repr(c) for c in row[:8]])
        return sheet_name, rows
    finally:
        wb.close()


def main() -> None:
    args = parse_args()

    orig_path = (ROOT / args.orig).resolve() if not args.orig.is_absolute() else args.orig
    arh_path = (ROOT / args.arh).resolve() if not args.arh.is_absolute() else args.arh

    print(f"ORIG: {orig_path.relative_to(ROOT) if orig_path.is_relative_to(ROOT) else orig_path}")
    print(f"ARH:  {arh_path.relative_to(ROOT) if arh_path.is_relative_to(ROOT) else arh_path}")

    orig_sheet, orig_rows = extract_rows(orig_path, args.rows)
    arh_sheet, arh_rows = extract_rows(arh_path, args.rows)

    print(f"\nFirst meaningful sheet ORIG: {orig_sheet!r}")
    print(f"First meaningful sheet ARH:  {arh_sheet!r}")

    print("\n" + "=" * 100)
    print("ORIG ROWS")
    print("=" * 100)
    for i, row in enumerate(orig_rows, start=1):
        joined = "  |  ".join(c[:50] for c in row[:7])
        print(f"  r{i:>3}: {joined}")

    print("\n" + "=" * 100)
    print("ARH ROWS")
    print("=" * 100)
    for i, row in enumerate(arh_rows, start=1):
        joined = "  |  ".join(c[:50] for c in row[:7])
        print(f"  r{i:>3}: {joined}")


if __name__ == "__main__":
    main()
