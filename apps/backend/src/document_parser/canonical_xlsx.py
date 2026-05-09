"""Block-oriented XLSX parser for Arhigon-style troskovnici.

Per Marko's guidance: Lexitor cares about *blocks of text* — one block =
one entity that the analyzer reasons about against ZJN/DKOM. We don't
need to model groups, sub-items, or unit-quantity columns precisely;
those details cost engineering effort but add little to the legal
analysis. We do still flag the math row(s) for a future deterministic
validator, but the primary product of the parser is the *merged text*
of each entity.

A block starts when:
    - column with R.b. (auto-detected) gets a new section/item label, OR
    - column with R.b. is empty for ≥ 2 consecutive rows after content
A block ends right before the next block starts. Pure header rows,
naslovnica/sadrzaj sheets and Numbers TOC sheets are skipped.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from src.document_parser.base import ParsedDocument, ParsedItem, ParserError

# ---------------------------------------------------------------------------
# Sheet classification

SKIP_SHEET_TOKENS = (
    "sadrz",  # sadržaj, sadrzaj
    "eksportiraj",  # Apple Numbers TOC export
    "summary",
    "export",
    "toc",
)
# Naslovne stranice — pojavljuju se u UI kao informativni "tekst", ali se
# NE analiziraju (nema math, nema brand-locka, nema DKOM provjera).
TEKST_SHEET_TOKENS = (
    "naslovn",  # naslovna, naslovnica, naslovni
    "nasl uk",  # naslovna ukratko (Arhigon abbreviation)
    "nasl",     # kratica koju koristi MILOS_Vinarija i sl.
    "title page",
    "cover",
    "korice",
)
OPCI_UVJETI_TOKENS = (
    "opci uvj",   # opći uvjeti (proper spelling, after diacritic strip)
    "opci uvij",  # opći uvijeti (common typo)
    "opci",       # kratica koju koristi MILOS_Vinarija i sl.
    "uvjeti",
    "uvijeti",
    "general cond",
)
REKAPITULACIJA_TOKENS = ("rekapitul", "recapitul", "summary of works")

# Croatian diacritics → ASCII so token matching is invariant to spelling.
# Lets "0_OPĆI UVIJETI" resolve to the same key as "opci uvijeti".
_DIACRITIC_TABLE = str.maketrans({"č": "c", "ć": "c", "ž": "z", "š": "s", "đ": "d"})


def _normalise(text: str) -> str:
    s = (text or "").lower().translate(_DIACRITIC_TABLE)
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def classify_sheet(name: str) -> str:
    norm = _normalise(name)
    if any(tok in norm for tok in SKIP_SHEET_TOKENS):
        return "skip"
    if any(tok in norm for tok in TEKST_SHEET_TOKENS):
        return "tekst"
    if any(tok in norm for tok in OPCI_UVJETI_TOKENS):
        return "opci_uvjeti"
    if any(tok in norm for tok in REKAPITULACIJA_TOKENS):
        return "rekapitulacija"
    return "stavke"


# ---------------------------------------------------------------------------
# Header / column detection

HEADER_TOKENS: dict[str, tuple[str, ...]] = {
    "rb": (
        "redni broj", "r b", "rb", "r br", "br", "broj",
        "stavka",  # Arhigon convention: column header "stavka" = position number
        "pozicija", "poz",
    ),
    "opis": (
        "opis stavke", "naziv stavke",
        "opis", "naziv", "predmet", "description",
        "stavka",  # fallback when there's no "opis stavke" column
    ),
    "jm": ("jedinica mjere", "jed mjere", "jed mj", "jm", "jedinica", "mjera", "unit"),
    "kol": ("kolicina stavke", "kolicina", "kol", "qty"),
    "cijena": ("jedinicna cijena", "jed cijena", "jedinicna", "cijena", "price"),
    "iznos": ("ukupna cijena", "ukupno", "iznos", "total", "vrijednost"),
}


@dataclass
class ColumnMapping:
    rb: int | None = None
    opis: int | None = None
    jm: int | None = None
    kol: int | None = None
    cijena: int | None = None
    iznos: int | None = None

    @property
    def is_minimally_complete(self) -> bool:
        return self.opis is not None


_TOKEN_RE_CACHE: dict[str, re.Pattern[str]] = {}


def _token_pattern(tok: str) -> re.Pattern[str]:
    """Word-boundary regex za header token. Cached jer se zove često.

    Substring match je nesiguran: token 'opis' bi se matched u 'propisima',
    'iznos' u 'iznositi', itd. Pravilo (column-name authority) traži da
    naziv stupca bude pravi naziv, ne dio veće riječi."""
    pat = _TOKEN_RE_CACHE.get(tok)
    if pat is None:
        pat = re.compile(rf"\b{re.escape(tok)}\b")
        _TOKEN_RE_CACHE[tok] = pat
    return pat


def _match_roles(text: str) -> list[tuple[str, int]]:
    """Return every (role, score) match for `text`. Score is the matched
    token length — longer match = more specific. Lets the caller pick
    intelligently when one cell matches multiple roles ("opis stavke"
    matches both "opis stavke" in opis and "stavka" in rb).

    Token mora biti cijela riječ — substring match unutar veće riječi
    ('opis' u 'propisima') NE računa se.

    Header label je kratak naslov tablice (max ~30 znakova), ne dugačka
    rečenica. Tekst u opisu stavke "armatura je predmet posebne stavke"
    sadrži header tokene (predmet, stavke) ali NIJE header — filtriramo
    cellove duže od ~40 znakova."""
    if not text or len(text) > 40:
        return []
    norm = _normalise(text)
    if not norm:
        return []
    out: list[tuple[str, int]] = []
    for role, tokens in HEADER_TOKENS.items():
        best = 0
        for tok in tokens:
            if _token_pattern(tok).search(norm) and len(tok) > best:
                best = len(tok)
        if best > 0:
            out.append((role, best))
    return out


def _match_role(text: str) -> str | None:
    """Backwards-compatible single-role lookup: pick the highest-scoring
    role for this text. Used by `detect_unit_quantity_swap` and similar
    secondary heuristics."""
    matches = _match_roles(text)
    if not matches:
        return None
    return max(matches, key=lambda x: x[1])[0]


def _row_to_strings(row: tuple[Cell, ...]) -> list[str]:
    return [
        ("" if c.value is None else str(c.value).replace("\n", " ").replace("\r", " "))
        for c in row
    ]


def _compute_column_stats(
    rows: list[tuple[Cell, ...]],
    skip_first: int = 0,
) -> tuple[dict[int, dict[str, int]], int]:
    """Walk the data rows and count per-column signals: section labels,
    long descriptive text, short unit-like tokens (jm), numerics, formulas.

    Used both by `_sniff_columns` (when there is no header at all) and by
    `find_header` (to content-infer roles like opis/rb/jm that the header
    row didn't explicitly name). `skip_first` is how many leading rows to
    skip — typically the header row itself."""
    if not rows:
        return {}, 0
    sample = rows[skip_first : skip_first + 100]
    max_cols = max((len(r) for r in sample), default=0)
    if max_cols == 0:
        return {}, 0
    stats: dict[int, dict[str, int]] = {
        c: {
            "section_labels": 0,
            "long_text": 0,
            "short_alpha": 0,
            "numeric": 0,
            "formula": 0,
            "non_empty": 0,
        }
        for c in range(max_cols)
    }
    # Jedinice mjere u Hrvatskoj mogu biti dulje od 5 znakova: "komplet"
    # (7), "izlazak" (7), "paušal" (6), "godina" (6). Dopuštamo do ~12
    # znakova ukupno — bilo što duže ide u long_text (>30) ili je
    # srednje-dugačka rečenica koja ne predstavlja jedinicu.
    short_alpha_re = re.compile(r"^[a-zA-Zčćžšđ][a-zA-Z0-9čćžšđ\s']{0,11}\.?$")
    for row in sample:
        for col_idx, cell in enumerate(row[:max_cols]):
            if cell is None or cell.value is None or cell.value == "":
                continue
            stats[col_idx]["non_empty"] += 1
            if cell.data_type == "f":
                stats[col_idx]["formula"] += 1
                continue
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                stats[col_idx]["numeric"] += 1
                continue
            s = str(v).strip()
            if not s:
                continue
            if _is_section_label(s):
                stats[col_idx]["section_labels"] += 1
                continue
            if len(s) > 30:
                stats[col_idx]["long_text"] += 1
            elif short_alpha_re.match(s):
                stats[col_idx]["short_alpha"] += 1
    return stats, max_cols


def _infer_missing_text_roles(
    mapping: ColumnMapping,
    stats: dict[int, dict[str, int]],
) -> ColumnMapping:
    """Content-inferira opis / rb / jm iz `stats` ne diraj postojeće
    explicit-header mappings. Koristi se kad `find_header` nađe header s
    nekoliko numerica (Količina/Jed.cijena/Ukupno) ali bez opisne tekstualne
    role, ili u sniff fallbacku.

    Pravila stupaca (vidi project_lexitor_column_rules.md):
    - opis = stupac s najdužim tekstom (long_text ≥ 2)
    - rb = stupac s najviše section labels (≥ 3)
    - jm = stupac s najviše short alfa tokena (≥ 2) — NE smije biti
      između cijena i iznos jer bi tu trebao biti "Jed. cijena"."""
    used: set[int] = {
        c for c in (
            mapping.rb, mapping.opis, mapping.jm,
            mapping.kol, mapping.cijena, mapping.iznos,
        ) if c is not None
    }

    def _pick(metric: str, threshold: int) -> int | None:
        candidates = [
            (c, st[metric]) for c, st in stats.items() if c not in used
        ]
        if not candidates:
            return None
        best = max(candidates, key=lambda x: x[1])
        return best[0] if best[1] >= threshold else None

    if mapping.opis is None:
        opis_col = _pick("long_text", 2)
        if opis_col is not None:
            mapping.opis = opis_col
            used.add(opis_col)

    if mapping.rb is None:
        rb_col = _pick("section_labels", 3)
        if rb_col is not None:
            mapping.rb = rb_col
            used.add(rb_col)

    if mapping.jm is None:
        jm_col = _pick("short_alpha", 2)
        if jm_col is not None:
            mapping.jm = jm_col
            used.add(jm_col)

    return mapping


def _sniff_columns(rows: list[tuple[Cell, ...]]) -> ColumnMapping:
    """Infer column→role mapping from data rows alone, for sheets that
    don't carry an explicit "stavka | opis | jed. mjere | …" header
    (common in older Croatian troskovnici).

    Returns an empty mapping when the data doesn't look structured
    enough — caller falls back to free-text parsing."""
    if not rows:
        return ColumnMapping()
    stats, _max_cols = _compute_column_stats(rows)
    if not stats:
        return ColumnMapping()

    mapping = ColumnMapping()
    mapping = _infer_missing_text_roles(mapping, stats)
    used: set[int] = {
        c for c in (mapping.rb, mapping.opis, mapping.jm) if c is not None
    }

    # Numeric columns get assigned to kol → cijena → iznos in the order
    # they appear left-to-right (typical Croatian layout). A column that
    # holds mostly formulas is most likely the iznos column (=kol×cijena).
    numeric_cols = sorted(
        (
            c for c, st in stats.items()
            if c not in used and (st["numeric"] + st["formula"]) >= 2
        )
    )
    if len(numeric_cols) >= 3:
        mapping.kol = numeric_cols[0]
        mapping.cijena = numeric_cols[1]
        mapping.iznos = numeric_cols[-1]
    elif len(numeric_cols) == 2:
        mapping.kol = numeric_cols[0]
        mapping.iznos = numeric_cols[1]
    elif len(numeric_cols) == 1:
        mapping.iznos = numeric_cols[0]

    if not mapping.is_minimally_complete:
        return ColumnMapping()
    return mapping


def find_header(rows: list[tuple[Cell, ...]]) -> tuple[int, ColumnMapping] | None:
    """Locate the header row and infer column→role mapping.

    The interesting case is ambiguity: a header like Arhigon's "stavka"
    can match both rb (position number) and opis (description). We
    handle it by collecting every match per column with its score, then
    assigning columns to roles greedily — most-specific match first —
    and falling back to a column's next-best role when its first choice
    is taken. As a final post-processing pass: if two columns matched
    cijena and iznos is still empty, the right-most one is promoted to
    iznos (Croatian convention "jedinična cijena" + "cijena (EUR)" =
    unit + total)."""
    for idx, row in enumerate(rows[:80]):
        cells = _row_to_strings(row)
        # Collect candidate roles per non-empty column
        candidates: list[tuple[int, list[tuple[str, int]]]] = []
        for col_idx, text in enumerate(cells):
            matches = _match_roles(text)
            if matches:
                candidates.append((col_idx, matches))
        if not candidates:
            continue

        # Sort columns by their best score DESC so the most-specific
        # header phrase ("opis stavke") gets first pick at its role.
        candidates.sort(key=lambda c: -max(s for _, s in c[1]))

        mapping = ColumnMapping()
        used_roles: set[str] = set()
        unassigned: list[int] = []
        for col_idx, matches in candidates:
            # Try roles in score-descending order until one is free
            for role, _score in sorted(matches, key=lambda m: -m[1]):
                if role not in used_roles and getattr(mapping, role) is None:
                    setattr(mapping, role, col_idx)
                    used_roles.add(role)
                    break
            else:
                unassigned.append(col_idx)

        # Post-process: if cijena got assigned but iznos didn't, and any
        # unassigned column also matched cijena tokens, that one is the
        # totals column ("cijena (EUR)" sitting next to "jedinična cijena").
        if mapping.cijena is not None and mapping.iznos is None:
            for col_idx in unassigned:
                text = cells[col_idx]
                if any(role == "cijena" for role, _ in _match_roles(text)):
                    if col_idx > mapping.cijena:
                        mapping.iznos = col_idx
                        used_roles.add("iznos")
                        break

        # Header je dovoljno kvalitetan ako pokriva ≥2 numerička/value
        # role (kol/cijena/iznos) ili ima opis. To znači: ako nadjemo
        # red s "Količina | Jedinična cijena | Ukupno" prihvaćamo ga
        # iako tu nema "Opis" (B5 je vjerojatno dynamic ref title).
        # Nedostajuće tekstualne role (opis/rb/jm) popunjavamo content
        # inference iz redova ISPOD headera — column-name authority:
        # stupac D je 'Količina' znači količina ide TAMO i nigdje
        # drugdje, čak i ako mu sniff misli da je "kol" stupac A.
        numeric_roles = sum(
            1 for r in ("kol", "cijena", "iznos") if getattr(mapping, r) is not None
        )
        accept = mapping.is_minimally_complete or numeric_roles >= 2
        if not accept:
            continue
        # Content-inferiraj opis/rb/jm samo kad nedostaju, koristeći
        # podatke od retka idx+1 nadalje (ne uključuje sam header).
        if mapping.opis is None or mapping.rb is None or mapping.jm is None:
            stats, _ = _compute_column_stats(rows, skip_first=idx + 1)
            if stats:
                mapping = _infer_missing_text_roles(mapping, stats)
        if mapping.is_minimally_complete:
            return idx, mapping

    # Token-based detection failed — sniff column roles from the data
    # itself. Older Croatian troskovnici (DV Netretić-style) often ship
    # without an explicit header row. Returns header_idx = -1 to signal
    # "no header to skip — start parsing from the very first row".
    sniffed = _sniff_columns(rows)
    if sniffed.is_minimally_complete:
        return -1, sniffed
    return None


# ---------------------------------------------------------------------------
# R.b. column fallback

SECTION_LABEL_RE = re.compile(
    r"^\s*(?:"
    r"[A-ZČĆŽŠĐ]\.?[\s\d.]*\d"                                   # A1, A.1, A.1.1, A 1
    r"|\d+(?:\.\d+){0,4}\.?"                                       # 1, 1., 1.1, 1.1.1
    r"|[A-ZČĆŽŠĐ]\.\d+(?:\.\d+){0,3}\.?"                          # A.1, A.1.1
    r"|[A-ZČĆŽŠĐ]\.[IVXLCM]+(?:\.\d+(?:\.\d+){0,3})?\.?"          # A.II, A.II.3, B.I.01, A.II.3.1
    r"|[A-ZČĆŽŠĐ]\.?"                                              # A, A. — single Croatian uppercase letter
    r"|[IVXLCM]+\.?"                                               # I, II, III, IV, V, VI, VII, VIII, IX, X (roman numerals)
    r")\s*$"
)
# Matches '1', '1.', '1.1', '1.1.1', 'A', 'A.', 'A.1', 'A.1.1', 'A.II', 'A.II.3',
# 'B.I.01', 'B.I.01.', '12', '23.4', 'I', 'II', 'III', 'IV', 'V', etc.

# Detection of UKUPNO/total label anywhere in the row text. Per Marko's
# guidance: an UKUPNO row carries the literal "UKUPNO" word (and usually
# the group label/name being summed) — distinct from stavka description
# rows. We use this together with the SUM formula detection so that even
# a hardcoded "UKUPNO" value (no formula) gets caught.
_UKUPNO_LABEL_RE = re.compile(r"\bUKUPNO\b|\bSUBTOTAL\b", re.IGNORECASE)


def _row_has_ukupno_label(
    row: tuple[Cell, ...],
    cache: dict[tuple[int, int], Any] | None = None,
) -> bool:
    """Return True if any cell in `row` contains the literal "UKUPNO" word."""
    for cell in row:
        if cell is None:
            continue
        val = _resolve(cell, cache)
        if isinstance(val, str) and _UKUPNO_LABEL_RE.search(val):
            return True
    return False


def _is_section_label(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    return bool(SECTION_LABEL_RE.match(s))


def _section_depth(label: str) -> int:
    """Heuristic depth ranking for a section label so we can manage a
    nesting stack: single Croatian uppercase letter (A, B, C…) sits at
    depth 1, roman numeral (I, II, III…) at depth 2, plain integer (1,
    2, …) at depth 3, dotted number (1.1, 1.1.1) at depth 4+. This
    mirrors how Croatian troskovnici nest groups → subgroups → stavke."""
    s = (label or "").strip().rstrip(".")
    if not s:
        return 0
    if re.fullmatch(r"\d+(?:\.\d+)+", s):
        return 3 + s.count(".")
    if re.fullmatch(r"\d+", s):
        return 3
    if re.fullmatch(r"[IVXLCM]+", s, re.IGNORECASE) and len(s) >= 1:
        # Roman numerals — but only when the value would actually parse.
        # "I", "II", etc. are common; "A" is a letter and won't reach this.
        return 2
    if re.fullmatch(r"[A-ZČĆŽŠĐ]", s):
        return 1
    if re.fullmatch(r"[A-ZČĆŽŠĐ]\.\d+(?:\.\d+)*", s):
        return 2 + s.count(".")
    # Fallback: treat unknown patterns as deepest so they don't pop ancestors
    return 99


def detect_rb_column(
    sample_rows: list[tuple[Cell, ...]], mapping: ColumnMapping
) -> ColumnMapping:
    if mapping.rb is not None:
        return mapping
    used = {mapping.opis, mapping.jm, mapping.kol, mapping.cijena, mapping.iznos}
    score: dict[int, int] = {}
    for row in sample_rows:
        for col_idx in range(min(len(row), 8)):
            if col_idx in used:
                continue
            cell = row[col_idx]
            if cell.value is None:
                continue
            if _is_section_label(cell.value):
                score[col_idx] = score.get(col_idx, 0) + 1
    if score:
        best_col, best_count = max(score.items(), key=lambda kv: kv[1])
        if best_count >= 2:
            mapping.rb = best_col
    return mapping


# ---------------------------------------------------------------------------
# Cell helpers

def _cell(row: tuple[Cell, ...], col: int | None) -> Cell | None:
    if col is None or col >= len(row):
        return None
    return row[col]


def _value(row: tuple[Cell, ...], col: int | None) -> Any:
    cell = _cell(row, col)
    return cell.value if cell is not None else None


def _resolve(cell: Cell | None, cache: dict[tuple[int, int], Any] | None) -> Any:
    """Cell value, but swap in the cached calculated value when the cell
    holds a formula. Lets `=B10` style cross-cell references in opis/label
    columns resolve to the human-readable text the user actually sees."""
    if cell is None:
        return None
    if cache is not None and cell.data_type == "f":
        return cache.get((cell.row, cell.column), cell.value)
    return cell.value


def _str(
    row: tuple[Cell, ...],
    col: int | None,
    cache: dict[tuple[int, int], Any] | None = None,
) -> str:
    cell = _cell(row, col)
    val = _resolve(cell, cache)
    if val is None:
        return ""
    return str(val).strip()


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", ".").strip()
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _has_math(
    row: tuple[Cell, ...],
    mapping: ColumnMapping,
    cache: dict[tuple[int, int], Any] | None = None,
) -> bool:
    """Math row indicators (column-name authority — vidi
    project_lexitor_column_rules.md). JM smije biti samo jedinica
    (string), a kol/cijena/iznos moraju biti **brojevi**. Cross-sheet
    text formula `=nasl!A2` koja vraća "VINSKI PODRUM …" NIJE iznos
    iako je formula — resolved value je string, ne broj."""
    # JM: bilo koji ne-prazan string indikator je dovoljan
    if mapping.jm is not None:
        jm_val = _resolve(_cell(row, mapping.jm), cache)
        if jm_val not in (None, ""):
            return True
    # Numerički stupci: resolved value MORA biti broj (ili numerički
    # string poput "3,50"). Formula koja vraća tekst se ne računa.
    for c in (mapping.kol, mapping.cijena, mapping.iznos):
        if c is None:
            continue
        v = _resolve(_cell(row, c), cache)
        if _num(v) is not None:
            return True
    return False


# Markers Marko uses to introduce a list of sub-positions inside an item.
_POSITIONS_HEADER_RE = re.compile(
    r"(?im)^(?:POZICIJE|POZICIJA|POPIS|SUBPOZICIJE|STAVKE|RAZRADA|RAZRAĐUJE)\s*:?\s*$"
)
_POSITION_BULLET_RE = re.compile(r"^\s*[-–•·]\s*(.+?)\s*$")

# UKUPNO / recap detection: a row whose iznos is a SUM(...) formula is
# never an actual stavka — it's either the stavka's own UKUPNO line (when
# the SUM covers rows belonging to that stavka) or a group/recap total
# (when the SUM covers rows from earlier blocks). We use the smallest
# referenced row to tell them apart.
_SUM_FUNC_RE = re.compile(r"\bSUM\s*\(\s*([^)]*)\s*\)", re.IGNORECASE)
_RANGE_REF_RE = re.compile(r"\$?[A-Z]+\$?(\d+)\s*:\s*\$?[A-Z]+\$?(\d+)")
_CELL_REF_RE = re.compile(r"\$?[A-Z]+\$?(\d+)")


def _extract_sum_rows(formula: Any) -> list[int] | None:
    """Return sorted unique row numbers that `formula` aggregates over,
    or None when it isn't an aggregation-style formula.

    Recognises:
    - `=SUM(F11:F43)` — range sum
    - `=SUM(F24,F36,F67)` — point-wise SUM with comma/semicolon list
    - `=F14+F19` — plain cell additions (no SUM function)
    - `=F42` — single-cell reference (often used for sub-UKUPNO that
      copies one stavka's total)

    Returns None for per-row math formulas like `=F42*E42` (kol×cijena
    product is NOT aggregating multiple rows, just one row's columns)
    so those don't get mis-classified as UKUPNO."""
    if not isinstance(formula, str) or not formula.lstrip().startswith("="):
        return None

    rows: set[int] = set()

    # SUM(...) function — handle first because the body inside may
    # contain colons/commas that the plain-arithmetic check rejects.
    m = _SUM_FUNC_RE.search(formula)
    if m is not None:
        body = m.group(1)
        for rm in _RANGE_REF_RE.finditer(body):
            a, b = int(rm.group(1)), int(rm.group(2))
            for r in range(min(a, b), max(a, b) + 1):
                rows.add(r)
        body_no_ranges = _RANGE_REF_RE.sub("", body)
        for rm in _CELL_REF_RE.finditer(body_no_ranges):
            rows.add(int(rm.group(1)))
        return sorted(rows) if rows else None

    # Plain cell additions: only +, -, $, digits, letters, whitespace
    # are allowed. * or / mean it's a product/division (per-row math),
    # not an aggregation.
    body = formula[1:].strip()
    if not body or "*" in body or "/" in body:
        return None
    if not re.fullmatch(r"[\s+\-$\d:A-Za-z]+", body):
        return None
    for rm in _RANGE_REF_RE.finditer(body):
        a, b = int(rm.group(1)), int(rm.group(2))
        for r in range(min(a, b), max(a, b) + 1):
            rows.add(r)
    body_no_ranges = _RANGE_REF_RE.sub("", body)
    for rm in _CELL_REF_RE.finditer(body_no_ranges):
        rows.add(int(rm.group(1)))
    return sorted(rows) if rows else None


def _row_sum_info(
    row: tuple[Cell, ...],
    iznos_col: int | None = None,
) -> tuple[list[int], str] | None:
    """Find a SUM-like formula on `row` and return (referenced rows,
    formula text).

    With `iznos_col` set (parse_stavke_sheet — header detection knew the
    iznos column), only that column is checked. This avoids false-flagging
    rows that have aggregation-style formulas in OTHER columns — e.g. a
    quantity cell `E46=E34+E39` summing prior rows is not an UKUPNO of
    F-column iznos.

    Without `iznos_col` (parse_rekapitulacija_sheet — no canonical mapping),
    every cell is scanned because rekapitulacija layouts vary."""
    if iznos_col is not None:
        cell = _cell(row, iznos_col)
        if cell is None or cell.value is None or cell.data_type != "f":
            return None
        rows = _extract_sum_rows(cell.value)
        if rows is None:
            return None
        return rows, str(cell.value)

    for cell in row:
        if cell is None or cell.value is None or cell.data_type != "f":
            continue
        rows = _extract_sum_rows(cell.value)
        if rows is not None:
            return rows, str(cell.value)
    return None


# Cross-sheet cell references in formulas. Two flavors are common:
#   'Sheet Name'!$F$46 / 'Sheet Name'!F46   — Excel quoted form
#   SheetName!F46                            — Excel bare (no spaces in name)
#   SheetName.F46                            — Apple Numbers / LibreOffice
# We tolerate `.` only when preceded by `=`/operator/`(`/`,`/`;`/space so it
# can't accidentally match a number literal like `1.5`.
_CROSS_SHEET_RE = re.compile(
    r"(?:'([^']+)'|(?<![A-Za-z0-9_])([A-Za-z_][\w]*))!\$?([A-Z]+)\$?(\d+)"
)
_APPLE_SHEET_RE = re.compile(
    r"(?:^=|[=+\-*/(),;\s])([A-Za-z_][\w]*)\.\$?([A-Z]+)\$?(\d+)"
)


def _extract_cross_sheet_refs(formula: Any) -> list[tuple[str, str, int]]:
    """Return list of (sheet_name, column_letter, row) for every cross-sheet
    cell reference in `formula`. Handles quoted Excel form, bare Excel form,
    and Apple Numbers `Sheet.Cell` form."""
    if not isinstance(formula, str) or not formula.lstrip().startswith("="):
        return []
    out: list[tuple[str, str, int]] = []
    for m in _CROSS_SHEET_RE.finditer(formula):
        sheet = (m.group(1) or m.group(2) or "").strip()
        col = m.group(3)
        row = int(m.group(4))
        if sheet:
            out.append((sheet, col, row))
    for m in _APPLE_SHEET_RE.finditer(formula):
        sheet = (m.group(1) or "").strip()
        col = m.group(2)
        row = int(m.group(3))
        if sheet:
            out.append((sheet, col, row))
    return out


def _row_first_formula(row: tuple[Cell, ...]) -> tuple[str, Any] | None:
    """Return (formula_text, calculated_value_or_None) for the first formula
    cell in `row`. Used in rekapitulacija parsing where the iznos column
    holds either a SUM or a cross-sheet reference."""
    for cell in row:
        if cell is None or cell.value is None or cell.data_type != "f":
            continue
        return str(cell.value), None
    return None


def _row_iznos_value(row: tuple[Cell, ...]) -> Any:
    """Return the rightmost numeric/formula value in `row`. Rekapitulacija
    layouts often skip the canonical column header detection, so we just
    take the last cell that holds something money-like."""
    for cell in reversed(row):
        if cell is None or cell.value in (None, ""):
            continue
        if cell.data_type == "f":
            return cell.value
        if isinstance(cell.value, (int, float)):
            return cell.value
    return None


def _extract_positions(text: str) -> list[str]:
    """Return the list of bullet labels under a 'POZICIJE:' header, if any."""
    if not text:
        return []
    lines = text.splitlines()
    started = False
    out: list[str] = []
    for line in lines:
        if not started:
            if _POSITIONS_HEADER_RE.match(line):
                started = True
            continue
        m = _POSITION_BULLET_RE.match(line)
        if m:
            label = m.group(1).strip()
            if label:
                out.append(label)
            continue
        # Empty line or a line that doesn't look like a bullet → end of list
        if line.strip() == "":
            if out:
                break
            continue
        break
    return out


# ---------------------------------------------------------------------------
# Block model

@dataclass
class _Block:
    label: str = ""
    title: str = ""
    title_row: int | None = None
    description_lines: list[tuple[int, str]] = field(default_factory=list)
    math_rows: list[dict[str, Any]] = field(default_factory=list)
    total_row: dict[str, Any] | None = None
    start_row: int = 0
    sheet: str = ""

    def add_description(self, row_idx: int, text: str) -> None:
        if text and text.strip():
            self.description_lines.append((row_idx, text))

    def add_math(
        self,
        row: tuple[Cell, ...],
        mapping: ColumnMapping,
        row_index: int,
        cache: dict[tuple[int, int], Any] | None = None,
        with_opis_label: bool = True,
    ) -> None:
        iznos_cell = _cell(row, mapping.iznos)
        kol_value = _resolve(_cell(row, mapping.kol), cache)
        cijena_value = _resolve(_cell(row, mapping.cijena), cache)
        is_formula = bool(iznos_cell and iznos_cell.data_type == "f")
        raw_iznos = (
            cache.get((iznos_cell.row, iznos_cell.column), iznos_cell.value)
            if cache is not None and iznos_cell is not None and iznos_cell.data_type == "f"
            else (iznos_cell.value if iznos_cell is not None else None)
        )
        # Always compute the total when both količina and jed.cijena are
        # numeric, so the UI can compare it to the Excel iznos and flag
        # discrepancies (Tier 1.4 — manipulacija jediničnim cijenama).
        computed_iznos: float | None = None
        kol_num = _num(kol_value)
        cijena_num = _num(cijena_value)
        if kol_num is not None and cijena_num is not None:
            computed_iznos = round(kol_num * cijena_num, 2)
        entry: dict[str, Any] = {
            "row": row_index,
            "jm": _str(row, mapping.jm, cache),
            "kol": kol_value,
            "cijena": cijena_value,
            "iznos": raw_iznos,
            "iznos_is_formula": is_formula,
            "computed_iznos": computed_iznos,
        }
        # Per Marko's rule: opis text on a math row IS its podstavka
        # label, not a separate body line. Section-header rows (where
        # opis is the stavka title, not a per-row label) skip this.
        if with_opis_label:
            opis_label = _str(row, mapping.opis, cache)
            if opis_label:
                entry["position_label"] = opis_label
        self.math_rows.append(entry)

    @property
    def is_empty(self) -> bool:
        return not (self.label or self.title or self.description_lines or self.math_rows)

    def merged_text(self) -> str:
        parts: list[str] = []
        if self.title:
            parts.append(self.title)
        parts.extend(text for _, text in self.description_lines)
        return "\n".join(p.strip() for p in parts if p and p.strip())


# ---------------------------------------------------------------------------
# Sheet parsers

EMPTY_BOUNDARY_LIMIT = 2  # 2+ consecutive empty rows close a block


def parse_stavke_sheet(
    ws,
    sheet_name: str,
    cache: dict[tuple[int, int], Any] | None = None,
) -> list[ParsedItem]:
    rows = list(ws.iter_rows())
    if not rows:
        return []

    header_info = find_header(rows)
    if header_info is None:
        return _raw_text_items(ws, sheet_name, cache, item_kind="opci_uvjeti")

    header_idx, mapping = header_info
    sample = rows[header_idx + 1 : header_idx + 41]
    mapping = detect_rb_column(sample, mapping)

    # Pre-header text rows (sheet title, opci uvjeti for the section,
    # specifications etc.) are emitted as opci_uvjeti items so they
    # aren't lost. Per Marko: A.I. Pripremni radovi typically has its
    # opći uvjeti at rows 1..36, then the header at row 37.
    # When header_idx is -1 (sniff fallback — no header row exists),
    # there's nothing pre-header to extract.
    preheader_items = (
        []
        if header_idx < 0
        else _extract_preheader_items(rows[:header_idx], sheet_name, cache)
    )

    blocks: list[_Block] = []
    group_sums: list[dict[str, Any]] = []
    current = _Block(sheet=sheet_name, start_row=header_idx + 2)
    empty_streak = 0

    for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        rb_str = _str(row, mapping.rb, cache)
        opis_str = _str(row, mapping.opis, cache)
        has_math = _has_math(row, mapping, cache)

        is_empty = not (rb_str or opis_str or has_math)
        is_new_section = bool(rb_str and _is_section_label(rb_str))

        if is_empty:
            empty_streak += 1
            if empty_streak >= EMPTY_BOUNDARY_LIMIT and not current.is_empty:
                blocks.append(current)
                current = _Block(sheet=sheet_name, start_row=offset)
            continue

        # Non-empty row resets streak
        empty_streak = 0

        # UKUPNO / recap intercept. Per Marko's rule, a row is an UKUPNO
        # if EITHER:
        #   (a) any cell holds an aggregation formula (SUM(...) or plain
        #       cell additions like =G42 / =G42+G53), or
        #   (b) any cell carries the literal "UKUPNO" / "UKP" / "Total" /
        #       "Suma" word — catches hardcoded UKUPNO rows that lack a
        #       SUM formula entirely.
        # Restrict SUM detection to the iznos column inside parse_stavke_sheet
        # so a quantity formula like E46=E34+E39 doesn't get misclassified
        # as an UKUPNO row.
        sum_info = _row_sum_info(row, iznos_col=mapping.iznos)
        ukupno_labelled = _row_has_ukupno_label(row, cache)
        # Suppress text-based UKUPNO trigger when the row also carries a
        # full kol×cijena pair: "ukupno površina ploča ophoda P=" is a
        # math row with descriptive text, not an aggregation. Only fire
        # text-based UKUPNO when at least one of kol/cijena is empty.
        if ukupno_labelled and not sum_info:
            kol_present = _num(_resolve(_cell(row, mapping.kol), cache)) is not None
            cijena_present = _num(_resolve(_cell(row, mapping.cijena), cache)) is not None
            if kol_present and cijena_present:
                ukupno_labelled = False
        if sum_info or ukupno_labelled:
            if sum_info:
                sum_rows, iznos_formula = sum_info
            else:
                # Text-only UKUPNO: read formula ONLY from the iznos
                # column. Any aggregation formula elsewhere in the row
                # (e.g. quantity sums) doesn't represent the UKUPNO total.
                iznos_cell = _cell(row, mapping.iznos)
                iznos_formula = (
                    iznos_cell.value
                    if iznos_cell is not None
                    and iznos_cell.data_type == "f"
                    and isinstance(iznos_cell.value, str)
                    else None
                )
                sum_rows = (
                    _extract_sum_rows(iznos_formula) or []
                    if iznos_formula
                    else []
                )
            sum_start = sum_rows[0] if sum_rows else offset
            label_parts = [
                _str(row, c, cache)
                for c in (mapping.opis, mapping.jm, mapping.kol, mapping.cijena, mapping.rb)
            ]
            label_text = " ".join(p for p in label_parts if p) or "UKUPNO"
            # Stavka-own vs group/recap classification:
            # - With sum_rows: smallest referenced row at/below current
            #   block's first row → stavka's own UKUPNO. Above → group sum.
            # - Without sum_rows (text-only UKUPNO with no formula or
            #   single ref): default to current stavka if non-empty.
            attaches_to_current = (
                not current.is_empty
                and (
                    (sum_rows and sum_start >= current.start_row)
                    or not sum_rows
                )
            )
            if attaches_to_current:
                # Stavka's own UKUPNO — attach to current block so we can
                # validate it later, but don't add it as a substavka.
                current.total_row = {
                    "row": offset,
                    "label": label_text,
                    "summed_rows": sum_rows,
                    "formula": iznos_formula,
                }
            else:
                # Group/recap sum — close current block, then collect this
                # sum so we can emit it as a separate "group_sum" item with
                # its own validation downstream.
                if not current.is_empty:
                    blocks.append(current)
                    current = _Block(sheet=sheet_name, start_row=offset)
                group_sums.append(
                    {
                        "row": offset,
                        "label": label_text,
                        "summed_rows": sum_rows,
                        "formula": iznos_formula,
                    }
                )
            continue

        if is_new_section:
            parent_rb = (current.label or "").rstrip(".")
            child_rb = rb_str.rstrip(".")

            # "Kompleti" pattern: a row with the SAME rb as the open
            # stavka — usually labelled "... - ukupno" — carries the
            # final kpl × cijena = iznos for the whole kit. Component
            # specification rows above had kol+jm only; this row brings
            # the price. Absorb it as the stavka's only math row,
            # don't open a new section.
            if (
                not current.is_empty
                and parent_rb
                and child_rb == parent_rb
                and has_math
            ):
                current.add_math(row, mapping, offset, cache, with_opis_label=False)
                continue

            # Child-numbered sections (e.g. "01.01.01" under "01.01" or
            # "1.1.1" under "1.1") that carry math values are podstavke
            # of the parent stavka, not new top-level stavke. Detect by
            # rb-prefix match and absorb as a math row of the current
            # block. The full rb (e.g. "01.01.01") is preserved as part
            # of the podstavka label so users see the numbering.
            #
            # IZNIMKA (Fix #3, 2026-05-10): apsorpcija samo ako parent već
            # ima math redove ili sam parent red (start_row) ima math. Kad
            # parent NEMA math (npr. "7.1 ZEMLJANI RADOVI" header bez
            # vrijednosti), to je section_header, a child "7.1.1" je
            # zasebna stavka — NE apsorbiraj. Ovaj _ARH layout (parent
            # section + djeca s math-om kao stavke) bio je pogrešno
            # spajan u 1 stavku s 5 podstavki umjesto 5 zasebnih stavki.
            parent_has_math = bool(current.math_rows)
            if (
                not current.is_empty
                and parent_rb
                and has_math
                and child_rb.startswith(parent_rb + ".")
                and parent_has_math
            ):
                current.add_math(row, mapping, offset, cache)
                last_math = current.math_rows[-1]
                opis_label = last_math.get("position_label", "")
                last_math["position_label"] = (
                    f"{rb_str} {opis_label}".strip() if opis_label else rb_str
                )
                continue

            if not current.is_empty:
                blocks.append(current)
            # Capture opis as the title initially. We only know whether
            # this section ends up being a stavka (has math) or a section
            # header (no math) once we see the rows that follow, so we
            # defer the decision: in the emit phase, stavke demote the
            # title to a description line (Marko's "treat stavka opis as
            # body text" rule), section headers keep it as a heading.
            current = _Block(
                label=rb_str,
                title=opis_str,
                title_row=offset if opis_str else None,
                sheet=sheet_name,
                start_row=offset,
            )
            if has_math:
                # Section row carries math → stavka. Don't read opis as
                # the math row's podstavka label since opis is the
                # stavka title (and will become its first description
                # line during emit).
                current.add_math(row, mapping, offset, cache, with_opis_label=False)
            continue

        # Continuation row inside the current block. Three sub-cases:
        # 1. Row has kol/jm but NO cijena AND NO iznos → "kompleti"
        #    component specification. The kit's price comes later on the
        #    same-rb totals row; this line just describes one part. We
        #    keep it as body text with a "kol jm" tail so the user sees
        #    "FID sklopka 4p, 25/0.03 A — 2 kom".
        # 2. Row has full math (cijena or iznos populated) → real math
        #    row. Opis on the same row becomes the podstavka label.
        # 3. Pure text row (no math at all) → body description.
        if has_math:
            cijena_present = (
                _resolve(_cell(row, mapping.cijena), cache) not in (None, "")
            )
            iznos_present = (
                _resolve(_cell(row, mapping.iznos), cache) not in (None, "")
            )
            if not cijena_present and not iznos_present:
                # Component spec (kompleti): keep as description
                kol_text = _str(row, mapping.kol, cache)
                jm_text = _str(row, mapping.jm, cache)
                tail = " ".join(p for p in (kol_text, jm_text) if p).strip()
                spec_text = opis_str
                if tail:
                    spec_text = f"{spec_text} — {tail}".strip(" —") if spec_text else tail
                if spec_text:
                    current.add_description(offset, spec_text)
                continue
            current.add_math(row, mapping, offset, cache)
        elif opis_str:
            current.add_description(offset, opis_str)

    if not current.is_empty:
        blocks.append(current)

    # Sheet-wide inventory of math rows (across all stavka blocks). Used
    # to validate group_sum coverage: for each group sum we want to know
    # which math rows fall in its row range and which are missing from
    # the sum's referenced rows.
    sheet_math_inventory: list[dict[str, Any]] = []
    math_inventory_by_row: dict[int, dict[str, Any]] = {}
    for block in blocks:
        for mr in block.math_rows:
            entry = {
                "row": mr.get("row"),
                "iznos": mr.get("iznos"),
                "computed_iznos": mr.get("computed_iznos"),
                "block_label": block.label,
                "block_title": block.title,
            }
            sheet_math_inventory.append(entry)
            if entry["row"] is not None:
                math_inventory_by_row[entry["row"]] = entry

    # Sheet-wide map of every total row (group sums + stavka own UKUPNO)
    # to its raw `summed_rows`. Lets us recursively expand a rollup sum
    # like "Ukupno A1+A2" (which references sub-totals, not math rows)
    # into the leaf math rows that ultimately contribute to it. Handles
    # arbitrary nesting depth via DFS with a visited set.
    total_rows_map: dict[int, list[int]] = {
        gs["row"]: gs["summed_rows"] for gs in group_sums
    }
    for block in blocks:
        if block.total_row:
            total_rows_map[block.total_row["row"]] = block.total_row["summed_rows"]
    math_row_set = set(math_inventory_by_row.keys())

    def _expand_to_leaves(seed_rows: list[int]) -> set[int]:
        """Recursively walk SUM references down to leaf math rows."""
        out: set[int] = set()
        visited: set[int] = set()
        stack: list[int] = list(seed_rows)
        while stack:
            r = stack.pop()
            if r in visited:
                continue
            visited.add(r)
            if r in math_row_set:
                out.add(r)
                continue
            sub = total_rows_map.get(r)
            if sub:
                stack.extend(sub)
        return out

    # Emit items in document order: stavka blocks and group_sums interleaved
    # by their first Excel row.
    ordered: list[tuple[int, str, Any]] = []
    for block in blocks:
        ordered.append((block.start_row, "stavka", block))
    for gs in group_sums:
        ordered.append((gs["row"], "group_sum", gs))
    ordered.sort(key=lambda x: x[0])

    # Stack of currently-open section headers, deepest last. Each stavka
    # inherits this stack as its `path` so the UI can show breadcrumbs and
    # the analyzer can reason about hierarchy ("which subgroup is this in").
    section_stack: list[dict[str, Any]] = []

    # Pre-header opci_uvjeti come first, before any actual stavke
    items: list[ParsedItem] = list(preheader_items)
    # Renumber positions to keep them contiguous as we append
    for i, it in enumerate(items):
        it.position = i

    for _row_idx, kind, payload in ordered:
        if kind == "group_sum":
            gs = payload
            summed = gs["summed_rows"] or []
            sum_set = set(summed)

            # Text-only UKUPNO with no formula at all (hardcoded value)
            # has empty summed_rows. We still emit it so the analyzer can
            # FAIL it as "iznos je upisan ručno", but skip range-based
            # bookkeeping that needs min()/max().
            if summed:
                sum_min, sum_max = min(summed), max(summed)
                effective = _expand_to_leaves(summed)
                is_rollup = bool(sum_set & set(total_rows_map.keys()))
                in_range_rows = {
                    mr["row"] for mr in sheet_math_inventory
                    if mr["row"] is not None and sum_min <= mr["row"] <= sum_max
                }
            else:
                effective = set()
                is_rollup = False
                in_range_rows = set()

            # Display rows: every math row this sum ultimately includes.
            # For leaf sums this is the directly-summed math rows. For
            # rollup sums it spans every leaf reached through sub-totals.
            covered_math = [
                math_inventory_by_row[r] for r in sorted(effective)
                if r in math_inventory_by_row
            ]

            missing_set = in_range_rows - effective - sum_set
            missing = [
                math_inventory_by_row[r] for r in sorted(missing_set)
                if r in math_inventory_by_row
            ]

            items.append(
                ParsedItem(
                    position=len(items),
                    label=(gs["label"] or "UKUPNO")[:200],
                    text=gs["label"] or "UKUPNO",
                    metadata={
                        "sheet": sheet_name,
                        "row": gs["row"],
                        "title_row": gs["row"],
                        "kind": "group_sum",
                        "summed_rows": summed,
                        "effective_summed_rows": sorted(effective),
                        "is_rollup": is_rollup,
                        "formula": gs["formula"],
                        "math_rows_in_range": covered_math,
                        "missing_rows": missing,
                    },
                )
            )
            continue

        block = payload
        text = block.merged_text()
        if not text and not block.math_rows:
            continue
        # Sidebar label: prefer a real heading; fall back to the first
        # descriptive line so the sidebar isn't just bare "4" for items
        # whose opis is a long paragraph.
        sidebar_label_pieces: list[str] = []
        if block.label:
            sidebar_label_pieces.append(block.label)
        if block.title:
            sidebar_label_pieces.append(block.title)
        elif block.description_lines:
            sidebar_label_pieces.append(block.description_lines[0][1])
        label = " · ".join(p for p in sidebar_label_pieces if p) or sheet_name

        # Block has descriptive content but no math rows → it's a
        # section/subgroup header or a "general terms" pre-amble, not an
        # actual stavka. Track its label on a stack so following stavke
        # know their hierarchical path; emit it as a slim section_header.
        if not block.math_rows:
            depth = _section_depth(block.label)
            while section_stack and section_stack[-1]["depth"] >= depth and depth > 0:
                section_stack.pop()
            # Cumulative full_path: ako stack ima parent, novi label se
            # prefiksa parent-ovim full_path-om (osim ako label već uključuje
            # parent). Npr. stack=[{label:XI., full_path:XI.}], block.label="1."
            # → full_path "XI.1.". Ovo pohranjeno u stack omogućuje
            # _build_full_rb-u da koristi puni hijerarhijski path bez
            # cumulative scan svaki put.
            local_lbl = (block.label or "").rstrip(".")
            if section_stack and local_lbl:
                parent_full = section_stack[-1].get("full_path", "").rstrip(".")
                if parent_full and not local_lbl.startswith(parent_full + "."):
                    full_path = f"{parent_full}.{local_lbl}."
                else:
                    full_path = f"{local_lbl}."
            else:
                full_path = f"{local_lbl}." if local_lbl else ""
            section_stack.append(
                {
                    "depth": depth,
                    "label": block.label,
                    "full_path": full_path,
                    "title": block.title,
                    "row": block.start_row,
                }
            )
            # If the block has any descriptive text under the heading
            # (NAPOMENA / opći uvjeti between subgroup and first stavka),
            # split it out: emit a slim section_header with just the
            # title, then emit each description line as its own
            # opci_uvjeti item so each row gets a row gutter and Lexitor
            # analiza panel.
            heading_label_parts = [
                p for p in (block.label, block.title) if p
            ]
            heading_text = " · ".join(heading_label_parts)
            if heading_text:
                items.append(
                    ParsedItem(
                        position=len(items),
                        label=heading_text[:200],
                        text=heading_text,
                        metadata={
                            "sheet": sheet_name,
                            "row": block.start_row,
                            "title_row": block.title_row,
                            "kind": "section_header",
                            "rb": block.label,
                            "title": block.title,
                            "depth": depth,
                            "path": [
                                {"label": s["label"], "title": s["title"]}
                                for s in section_stack
                            ],
                        },
                    )
                )
            for row_idx, line in block.description_lines:
                items.append(
                    ParsedItem(
                        position=len(items),
                        label=(line if len(line) <= 200 else line[:197] + "…"),
                        text=line,
                        metadata={
                            "sheet": sheet_name,
                            "row": row_idx,
                            "kind": "opci_uvjeti",
                            "path": [
                                {"label": s["label"], "title": s["title"]}
                                for s in section_stack
                            ],
                        },
                    )
                )
            continue

        # Stavka: demote any title to the first description line so the
        # card body shows the opis paragraph (with a row gutter) and the
        # h2 heading is suppressed. Marko's rule: stavka opis cells are
        # body text, never an extracted title.
        if block.title and block.title_row is not None:
            block.description_lines.insert(
                0, (block.title_row, block.title)
            )
            block.title = ""
            block.title_row = None
        # Re-derive merged text after the demote so text_rows + text
        # stay consistent.
        text = block.merged_text()

        positions = _extract_positions(text)
        # Pair positions with math rows whenever both lists are present
        # and have matching cardinalities. This is Marko's standard
        # convention: "POZICIJE: - nastamba - bazeni …" + N math rows.
        if positions and block.math_rows and len(positions) == len(block.math_rows):
            for label_text, math_row in zip(positions, block.math_rows, strict=True):
                math_row["position_label"] = label_text

        items.append(
            ParsedItem(
                position=len(items),
                label=label[:200],
                text=text,
                metadata={
                    "sheet": sheet_name,
                    "row": block.start_row,
                    "title_row": block.title_row,
                    "text_rows": [
                        {"row": r, "text": t}
                        for r, t in block.description_lines
                    ],
                    "kind": "stavka",
                    # Full hijerarhijski rb po _ARH konvenciji ("2.7.01."
                    # umjesto samo "01."). UI prikazuje full path; analyzer
                    # ga koristi za rb pairing protiv ground truth-a.
                    "rb": _build_full_rb(section_stack, block.label),
                    "rb_local": block.label,
                    "title": block.title,
                    "math_rows": block.math_rows,
                    "total_row": block.total_row,
                    "positions": positions or None,
                    "path": [
                        {"label": s["label"], "title": s["title"]}
                        for s in section_stack
                    ],
                },
            )
        )
    return items


def _build_full_rb(stack: list[dict[str, Any]], local: str) -> str:
    """Spoji section stack + lokalnu stavku rb u **puni hijerarhijski path**.

    _ARH konvencija (vidi project_lexitor_troskovnik_format.md):
        section_header A.        ← stack push
        section_header A.I.      ← stack push
        section_header A.I.1.    ← stack push
        stavka         A.I.1.1.  ← rb je full path

    Pravila (revizija 2026-05-10 nakon DV Netretić outlier):

    1. Ako local rb već ima ≥2 segmenta (npr. "XI.1.1.", "1.3.2."),
       pretpostavi da je već full path → vrati as-is. Naručitelji koji
       koriste full-path Excel layout (CONVEXO/Netretić tip) dobivaju
       lokalni rb koji već reflektira hijerarhiju.

    2. Ako local rb je jedan segment ("01.", "1.", "A."), prepend
       zadnjeg section labela:
           stack=[XI.], local=1. → XI.1.
           stack=[2.7.], local=01. → 2.7.01.
       Stack-ovi gdje su elementi već full path (XI., XI.1., XI.1.1.)
       koristimo SAMO zadnji element kao prefix.
    """
    if not stack or not local:
        return local
    # Koristimo cumulative full_path ako je dostupan (push gradi cumulative
    # path: stack=[XI.→full_path=XI., 1.→full_path=XI.1.]). Fallback na
    # label za stare push-eve bez full_path.
    top = stack[-1] if stack else {}
    parent_full = top.get("full_path") or top.get("label", "")
    parent_clean = parent_full.rstrip(".")
    local_clean = local.rstrip(".")
    if not parent_clean or not local_clean:
        return local

    # Idempotentno: ako local već počinje s parent full_path-om, ne prependa-ti.
    #   stack[-1].full_path=XI.1., local=XI.1.1. → već je full → "XI.1.1."
    if local_clean == parent_clean or local_clean.startswith(parent_clean + "."):
        return local

    # Overlap detection: ako parent zadnji segment matchira local prvi
    # segment, drop dupliranje. CONVEXO primjer:
    #   parent.full_path = "XI.1." (XI + section 1.)
    #   local = "1.1."  (lokalni rb stavke u section 1.)
    #   bez overlap-a: "XI.1.1.1." — KRIVO (4 segmenta)
    #   s overlap-om: "XI.1.1."     — TOČNO
    parent_segs = [s for s in parent_clean.split(".") if s]
    local_segs = [s for s in local_clean.split(".") if s]
    if parent_segs and local_segs and parent_segs[-1] == local_segs[0]:
        merged_segs = parent_segs + local_segs[1:]
    else:
        merged_segs = parent_segs + local_segs
    if not merged_segs:
        return local
    return ".".join(merged_segs) + "."


_RECAP_GRAND_RE = re.compile(r"\bSVEUKUPN|GRAND\s*TOTAL|TOTAL\s*INC\b", re.IGNORECASE)
_RECAP_PDV_RE = re.compile(r"\bPDV\b|\bVAT\b", re.IGNORECASE)
_RECAP_TOTAL_RE = re.compile(r"\bUKUPNO\b|\bSUBTOTAL\b", re.IGNORECASE)
_GROUP_LETTER_RE = re.compile(r"^[A-ZČĆŽŠĐ]\.?$")
_ROMAN_NUM_RE = re.compile(r"^[IVXLCM]+\.?$", re.IGNORECASE)


def parse_rekapitulacija_sheet(
    ws,
    sheet_name: str,
    cache: dict[tuple[int, int], Any] | None = None,
) -> list[ParsedItem]:
    """Structured parse of a Rekapitulacija sheet.

    Each non-empty row becomes one of:
    - recap_section:   group header (A/B/C... + group name, no money column)
    - recap_line:      references another sheet's total via cross-sheet ref
                       OR carries an iznos value pointing at a group/sheet
    - recap_subtotal:  SUM of preceding recap rows (group's own UKUPNO)
    - recap_total:     grand UKUPNO of the rekapitulacija
    - recap_pdv:       PDV/VAT calculation
    - recap_grand:     SVEUKUPNO (UKUPNO + PDV)
    - recap_extra:     hardcoded line with a number but no formula (e.g.
                       "Nepredviđeni radovi 5%")
    """
    items: list[ParsedItem] = []
    rows = list(ws.iter_rows())
    if not rows:
        return items

    section_header: str | None = None  # latest A/B/C group label

    for offset, row in enumerate(rows, start=1):
        cell_strings = []
        cell_kinds: list[str] = []  # parallel to cell_strings: "text" / "number" / "formula" / ""
        for c in row:
            if c is None:
                cell_strings.append("")
                cell_kinds.append("")
                continue
            raw_value = c.value
            resolved = _resolve(c, cache)
            if resolved in (None, ""):
                cell_strings.append("")
                cell_kinds.append("")
                continue
            text = str(resolved).strip()
            cell_strings.append(text)
            if isinstance(resolved, (int, float)) and not isinstance(resolved, bool):
                cell_kinds.append("number")
            elif (
                isinstance(raw_value, str)
                and raw_value.startswith("=")
                and not isinstance(resolved, (int, float))
            ):
                cell_kinds.append("formula")
            else:
                cell_kinds.append("text")
        if not any(cell_strings):
            continue

        # Sheet-title row (single big descriptive cell, no money / formula)
        sum_info = _row_sum_info(row)
        # In rekapitulacija every "iznos" line typically holds THREE
        # cross-sheet formulas — one each for rb, opis and the actual
        # iznos column. Validation cares about the iznos formula (the
        # one that returns a number) — point at the source sheet's
        # UKUPNO. Pick the right-most cell whose resolved value is
        # numeric; if none qualify, fall back to the right-most formula.
        formula_cells: list[Cell] = [
            c for c in row
            if c is not None and c.value is not None and c.data_type == "f"
        ]
        iznos_formula_cell: Cell | None = None
        for c in reversed(formula_cells):
            resolved = _resolve(c, cache)
            if isinstance(resolved, (int, float)) and not isinstance(resolved, bool):
                iznos_formula_cell = c
                break
        if iznos_formula_cell is None and formula_cells:
            iznos_formula_cell = formula_cells[-1]
        formula = (
            str(iznos_formula_cell.value) if iznos_formula_cell is not None else None
        )
        cross_refs = _extract_cross_sheet_refs(formula or "")
        iznos_value = _row_iznos_value(row)

        # Free-text label of this row: keep only descriptive (non-numeric)
        # strings. A resolved cross-sheet reference like =armiracki!F26 in
        # the iznos column comes back as a number (101050.4) — that's the
        # iznos, not part of the title. Don't concatenate it into the label.
        text_parts = [
            s
            for s, k in zip(cell_strings, cell_kinds, strict=True)
            if s and k == "text"
        ]
        full_text = " ".join(text_parts).strip()

        # Heuristic: section header — one of the cells is a single uppercase
        # letter (group marker A/B/C…) AND no formula/iznos in the row.
        first_short = next((s for s in cell_strings if s), "")
        is_letter_header = bool(_GROUP_LETTER_RE.match(first_short))
        if (
            is_letter_header
            and not sum_info
            and not formula
            and (iznos_value is None or full_text)
        ):
            # A standalone group header: "A   GRAĐEVINSKO-OBRTNIČKI RADOVI"
            section_header = full_text or first_short
            items.append(
                ParsedItem(
                    position=len(items),
                    label=full_text[:200] or first_short,
                    text=full_text or first_short,
                    metadata={
                        "sheet": sheet_name,
                        "row": offset,
                        "kind": "recap_section",
                        "section": section_header,
                    },
                )
            )
            continue

        # Classify by content priority: SVEUKUPNO > PDV > UKUPNO/SUM > line.
        is_grand = bool(_RECAP_GRAND_RE.search(full_text))
        is_pdv = (not is_grand) and bool(_RECAP_PDV_RE.search(full_text))
        is_ukupno = (
            not is_grand and not is_pdv and bool(_RECAP_TOTAL_RE.search(full_text))
        )

        if is_grand:
            kind = "recap_grand"
        elif is_pdv:
            kind = "recap_pdv"
        elif sum_info or is_ukupno:
            kind = "recap_subtotal" if section_header else "recap_total"
        elif cross_refs:
            kind = "recap_line"
        elif iznos_value is not None:
            # Hardcoded number with descriptive text — additional cost line
            kind = "recap_extra"
        else:
            # Free text with no number — likely a continuation/description
            kind = "recap_section"

        # Title row at the very top (no group context yet, just descriptive text)
        if kind == "recap_section" and not section_header and not is_letter_header:
            # Skip the sheet's title row from generating an item.
            continue

        # Resolve summed rows for SUM formulas
        summed_rows: list[int] = sum_info[0] if sum_info else []
        sum_formula = sum_info[1] if sum_info else None

        # Best-effort label
        label = full_text or section_header or kind.upper()

        items.append(
            ParsedItem(
                position=len(items),
                label=label[:200],
                text=full_text or label,
                metadata={
                    "sheet": sheet_name,
                    "row": offset,
                    "title_row": offset,
                    "kind": kind,
                    "section": section_header,
                    "iznos": iznos_value if not isinstance(iznos_value, str) else None,
                    "iznos_raw": iznos_value,
                    "formula": formula or sum_formula,
                    "cross_sheet_refs": [
                        {"sheet": s, "col": c, "row": r}
                        for (s, c, r) in cross_refs
                    ],
                    "summed_rows": summed_rows or None,
                },
            )
        )

    return items


def _raw_text_items(
    ws,
    sheet_name: str,
    cache: dict[tuple[int, int], Any] | None = None,
    item_kind: str = "opci_uvjeti",
) -> list[ParsedItem]:
    """Fallback for sheets without a recognisable header — every non-empty
    row becomes a standalone item. Used for opci_uvjeti and any
    unrecognised sheet structure. Resolves cross-sheet formula references
    via cache so labels like `=nasl uk!C1` show as the actual investor
    name instead of raw formula text."""
    items: list[ParsedItem] = []
    for idx, row in enumerate(ws.iter_rows(), start=1):
        cells: list[str] = []
        for c in row:
            if c is None:
                continue
            val = _resolve(c, cache)
            if val in (None, ""):
                continue
            s = str(val).strip()
            if s:
                cells.append(s)
        if not cells:
            continue
        text = " ".join(cells)
        # Use the row text itself as the item label so the sidebar shows
        # something readable. Truncate to keep it tidy; full text stays
        # in `text`. Avoid the old "sheet · red N" pattern — splitLabel
        # would parse "red N" as a title, so cards displayed "red 1" as
        # the heading.
        label = text if len(text) <= 200 else text[:197] + "…"
        items.append(
            ParsedItem(
                position=len(items),
                label=label,
                text=text,
                metadata={"sheet": sheet_name, "row": idx, "kind": item_kind},
            )
        )
    return items


def _extract_preheader_items(
    pre_rows: list[tuple[Cell, ...]],
    sheet_name: str,
    cache: dict[tuple[int, int], Any] | None = None,
) -> list[ParsedItem]:
    """Pull text-only rows from the area above the header row in a stavke
    sheet and emit them as opci_uvjeti items. Captures the general-terms
    paragraph that often sits at the top of a section (e.g. rows 1–36 in
    'A.I. PRIPREMNI RADOVI' before the actual header at row 37)."""
    items: list[ParsedItem] = []
    for idx, row in enumerate(pre_rows, start=1):
        cells: list[str] = []
        for c in row:
            if c is None:
                continue
            val = _resolve(c, cache)
            if val in (None, ""):
                continue
            s = str(val).strip()
            if s:
                cells.append(s)
        if not cells:
            continue
        text = " ".join(cells)
        label = text if len(text) <= 200 else text[:197] + "…"
        items.append(
            ParsedItem(
                position=len(items),
                label=label,
                text=text,
                metadata={"sheet": sheet_name, "row": idx, "kind": "opci_uvjeti"},
            )
        )
    return items


# ---------------------------------------------------------------------------
# Public entry

def _normalise_sheet_name(name: str) -> str:
    """Case-insensitive, whitespace-collapsed key for sheet lookup. Excel
    formula `=Pripremni!F35` should resolve to a sheet named `pripremni`
    or `PRIPREMNI` regardless of case (Apple Numbers tends to lowercase)."""
    return " ".join(str(name or "").strip().lower().split())


def _build_document_registry(items: list[ParsedItem]) -> dict[str, dict[str, Any]]:
    """Sheet → {ukupno_rows, math_rows, grand_ukupno_rows} from parsed items.

    Lets us validate a recap_line's `=zemljani!F46` reference: does row 46
    in sheet "zemljani" actually correspond to that sheet's UKUPNO, or
    is it pointing at some random math row mid-document?"""
    registry: dict[str, dict[str, Any]] = {}

    def _ensure(sheet_name: str) -> dict[str, Any]:
        key = _normalise_sheet_name(sheet_name)
        if key not in registry:
            registry[key] = {
                "name": sheet_name,
                "ukupno_rows": {},
                "math_rows": set(),
                "grand_ukupno_rows": set(),
            }
        return registry[key]

    for item in items:
        meta = item.metadata or {}
        sheet = meta.get("sheet")
        if not sheet:
            continue
        kind = meta.get("kind", "")
        entry = _ensure(sheet)

        if kind == "group_sum":
            row = meta.get("row")
            summed = meta.get("summed_rows") or []
            if row is not None:
                entry["ukupno_rows"][row] = {
                    "row": row,
                    "summed_rows": list(summed),
                    "label": item.text or item.label,
                    "is_rollup": bool(meta.get("is_rollup")),
                }
        elif kind == "stavka":
            for mr in meta.get("math_rows") or []:
                mr_row = mr.get("row")
                if mr_row is not None:
                    entry["math_rows"].add(mr_row)
            total_row = meta.get("total_row")
            if total_row and total_row.get("row") is not None:
                entry["ukupno_rows"][total_row["row"]] = {
                    "row": total_row["row"],
                    "summed_rows": list(total_row.get("summed_rows") or []),
                    "label": total_row.get("label", "UKUPNO"),
                    "is_rollup": False,
                }
        elif kind in ("recap_subtotal", "recap_total", "recap_grand"):
            row = meta.get("row")
            summed = meta.get("summed_rows") or []
            if row is not None:
                entry["ukupno_rows"][row] = {
                    "row": row,
                    "summed_rows": list(summed),
                    "label": item.text or item.label,
                    "is_rollup": False,
                }
                if kind == "recap_grand":
                    entry["grand_ukupno_rows"].add(row)

    # Mark the widest UKUPNO of each sheet as that sheet's grand UKUPNO so
    # cross-sheet refs from rekapitulacija can be checked against the
    # right target. "Widest" = greatest row-range coverage of summed_rows.
    for entry in registry.values():
        ukupno_rows = entry["ukupno_rows"]
        if not ukupno_rows:
            continue

        def _coverage(u: dict[str, Any]) -> int:
            sr = u.get("summed_rows") or []
            return (max(sr) - min(sr) + 1) if sr else 0

        widest = max(ukupno_rows.values(), key=_coverage, default=None)
        if widest is not None and _coverage(widest) > 0:
            entry["grand_ukupno_rows"].add(widest["row"])

    return registry


def _validate_cross_sheet_ref(
    ref: dict[str, Any],
    registry: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Classify a recap_line's cross-sheet reference against the document
    registry. Returns a copy of `ref` enriched with status/kind/message."""
    sheet_key = _normalise_sheet_name(ref.get("sheet", ""))
    target_row = ref.get("row")

    if sheet_key not in registry:
        return {
            **ref,
            "validation_status": "fail",
            "validation_kind": "missing_sheet",
            "message": f"Sheet '{ref.get('sheet')}' ne postoji u dokumentu.",
        }
    entry = registry[sheet_key]
    if target_row in entry["grand_ukupno_rows"]:
        u = entry["ukupno_rows"].get(target_row, {})
        return {
            **ref,
            "validation_status": "ok",
            "validation_kind": "grand_ukupno",
            "message": (
                f"Pokazuje na grand UKUPNO sheeta '{entry['name']}' "
                f"(red {target_row})."
            ),
            "target_label": u.get("label"),
        }
    if target_row in entry["ukupno_rows"]:
        u = entry["ukupno_rows"][target_row]
        return {
            **ref,
            "validation_status": "ok",
            "validation_kind": "sub_ukupno",
            "message": (
                f"Pokazuje na podgrupni UKUPNO sheeta '{entry['name']}' "
                f"(red {target_row}: {u.get('label', 'UKUPNO')})."
            ),
            "target_label": u.get("label"),
        }
    if target_row in entry["math_rows"]:
        return {
            **ref,
            "validation_status": "fail",
            "validation_kind": "math_row",
            "message": (
                f"Pokazuje na pojedinačnu math stavku u retku {target_row} "
                f"sheeta '{entry['name']}' — trebao bi pokazivati na UKUPNO."
            ),
        }
    return {
        **ref,
        "validation_status": "warn",
        "validation_kind": "unknown",
        "message": (
            f"Referenca na red {target_row} u sheetu '{entry['name']}' "
            f"nije prepoznata ni kao stavka ni kao UKUPNO."
        ),
    }


def _build_formula_cache(cached_ws) -> dict[tuple[int, int], Any]:
    """Build a (row, col) -> calculated value map for an entire sheet so
    label cells holding `=B10` style references can show what the user
    actually sees instead of the raw formula text."""
    cache: dict[tuple[int, int], Any] = {}
    for row in cached_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cache[(cell.row, cell.column)] = cell.value
    return cache


def parse_canonical_xlsx(path: Path) -> ParsedDocument:
    try:
        wb = load_workbook(filename=str(path), data_only=False, read_only=False)
        wb_cached = load_workbook(filename=str(path), data_only=True, read_only=False)
    except InvalidFileException as exc:
        raise ParserError(f"Nevažeći XLSX: {exc}") from exc

    items: list[ParsedItem] = []
    sheets_meta: list[dict[str, Any]] = []

    try:
        for ws in wb.worksheets:
            kind = classify_sheet(ws.title)
            cached_ws = (
                wb_cached[ws.title] if ws.title in wb_cached.sheetnames else None
            )
            cache = _build_formula_cache(cached_ws) if cached_ws is not None else {}

            if kind == "skip":
                sheet_items: list[ParsedItem] = []
            elif kind == "stavke":
                sheet_items = parse_stavke_sheet(ws, ws.title, cache)
            elif kind == "rekapitulacija":
                sheet_items = parse_rekapitulacija_sheet(ws, ws.title, cache)
            elif kind == "tekst":
                # Naslovna stranica — vidljivo u UI kao info, ali ne
                # ulazi u nikakve provjere (nema brand-lock ni mock).
                sheet_items = _raw_text_items(ws, ws.title, cache, item_kind="tekst")
            else:
                # opci_uvjeti — free-text per row, no structured math
                sheet_items = _raw_text_items(ws, ws.title, cache)
                for it in sheet_items:
                    it.metadata["kind"] = kind

            for item in sheet_items:
                item.position = len(items)
                items.append(item)

            sheets_meta.append({"name": ws.title, "kind": kind, "items": len(sheet_items)})
    finally:
        wb.close()
        wb_cached.close()

    # Document-wide registry of UKUPNO rows + math rows. Used to validate
    # cross-sheet references in rekapitulacija lines: a recap_line's
    # `=zemljani!F46` should point at zemljani's grand UKUPNO row, not at
    # some random math row.
    registry = _build_document_registry(items)
    for item in items:
        meta = item.metadata or {}
        if meta.get("kind") != "recap_line":
            continue
        refs = meta.get("cross_sheet_refs") or []
        if not refs:
            continue
        meta["ref_validation"] = [
            _validate_cross_sheet_ref(ref, registry) for ref in refs
        ]

    return ParsedDocument(
        items=items,
        metadata={"format": "xlsx", "parser": "canonical-block", "sheets": sheets_meta},
    )
