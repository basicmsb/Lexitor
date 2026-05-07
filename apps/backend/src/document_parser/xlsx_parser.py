from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.document_parser.base import ParsedDocument, ParsedItem, ParserError


def parse_xlsx(path: Path) -> ParsedDocument:
    items: list[ParsedItem] = []
    sheets_meta: list[dict[str, object]] = []
    try:
        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    except InvalidFileException as exc:
        raise ParserError(f"Nevažeći XLSX: {exc}") from exc

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            row_count = 0
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if not cells:
                    continue
                row_count += 1
                items.append(
                    ParsedItem(
                        position=len(items),
                        text=" | ".join(cells),
                        label=f"List „{sheet_name}”, red {row_index}",
                        metadata={
                            "sheet": sheet_name,
                            "row": row_index,
                            "source": "xlsx",
                        },
                    )
                )
            sheets_meta.append({"name": sheet_name, "rows": row_count})
    finally:
        workbook.close()

    return ParsedDocument(
        items=items,
        metadata={"format": "xlsx", "sheets": sheets_meta},
    )
