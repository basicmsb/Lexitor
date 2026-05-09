"""Strogo-deterministički parser za Arhigon-ove _ARHIGON_R0x.xlsx fajlove.

Ovi fajlovi su standardizirana ground-truth verzija originalnog troškovnika
(Marko ih ručno prepravi po fiksnim pravilima):

  - A: rb (numeracija) — uvijek u stupcu A, nikad u više stupaca
  - B: opis stavke + opis podstavke
  - C: jed. mjere
  - D: količina
  - E: jedinična cijena
  - F: iznos (najčešće formula =D*E)
  - J: section markers unutar stavke:
      "sn" — Section name (kratki naziv prije više math redova)
      "st" — Section text (dulji tekst između podstavki)
      "ss" — Section sum (suma sekcije, najčešće u elektro)
  - Bez praznih redova između dijelova jedne stavke (header→...→sljedeća stavka)

Ovaj parser **služi kao "prevoditelj"** — daje ground-truth strukturu protiv
koje testiramo originalni `canonical_xlsx` parser. Razlike (diff) postaju
input za nova pravila u original parseru.

Vidi `project_lexitor_troskovnik_format.md` u memoriji za kontekst."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from src.document_parser.base import ParsedDocument, ParsedItem, ParserError
from src.document_parser.canonical_xlsx import (
    _build_formula_cache,
    _normalise,
    _resolve,
    classify_sheet,
)

# Strogi mapping — _ARH konvencija
_COL_RB = 0      # A
_COL_OPIS = 1    # B
_COL_JM = 2      # C
_COL_KOL = 3     # D
_COL_CIJENA = 4  # E
_COL_IZNOS = 5   # F
_COL_J = 9       # J — section markers

_SECTION_MARKERS = {"sn", "st", "ss"}

# Pattern za rb (1.1., 1.1.01., A.II., 01., …)
_RB_RE = re.compile(
    r"^[A-ZČĆŽŠĐ0-9]+(?:\.[A-ZČĆŽŠĐ0-9]+)*\.?$"
)


@dataclass
class _Block:
    """Jedan blok = jedna stavka (rb + opis + math redovi + section markers)."""
    sheet: str
    rb: str = ""
    title: str = ""
    title_row: int | None = None
    description_lines: list[tuple[int, str]] = field(default_factory=list)
    math_rows: list[dict[str, Any]] = field(default_factory=list)
    section_markers: list[dict[str, Any]] = field(default_factory=list)
    start_row: int = 0

    @property
    def is_empty(self) -> bool:
        return not (
            self.rb or self.title or self.description_lines
            or self.math_rows or self.section_markers
        )


def _cell_value(row: tuple[Cell, ...], col: int, cache: dict) -> Any:
    """Resolved value (formula → cached) — jednako kao u canonical_xlsx."""
    if col >= len(row):
        return None
    return _resolve(row[col], cache)


def _str(row: tuple[Cell, ...], col: int, cache: dict) -> str:
    v = _cell_value(row, col, cache)
    return "" if v is None else str(v).strip()


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", ".").strip()
    if not s or s.startswith("="):
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _is_rb(text: str) -> bool:
    """Je li ovo legitiman rb? Pravila iz _ARH:
    - "1.", "1.1.", "1.1.01." — dotted numerics
    - "A.", "B.I.", "I.1." — slovima/rimskim
    - mora završiti točkom (po _ARH konvenciji)"""
    if not text:
        return False
    return bool(_RB_RE.match(text))


def _detect_layout(rows: list[tuple[Cell, ...]], cache: dict) -> str:
    """Razlikuje Layout A (Marko-ov standard: A=rb, B=opis, C-F=math) od
    Layout B (stariji predlošci: A=section formula, B=dinamic formula,
    C=rb, D=opis, E-H=math). Layout B susrećemo u Westin/Abilia/EmiLu/EKP.

    Vraća "A" ili "B". Default "A" ako nije jasno.

    Heuristike, redom:
    1. Header match: "opis"/"naziv stavke" u B → A, u D → B
    2. Layout B telltale: B-stupac sadrži kompleksnu formulu na većini
       redova (npr. `=IF(C3="","", A3&IF(...`). U Layout A B-stupac je
       slobodan tekst opisa, nikad formula koja gleda druge stupce.
    """
    # 1. Header detekcija
    for row in rows[:10]:
        for col, layout in ((_COL_OPIS, "A"), (3, "B")):
            val = _str(row, col, cache)
            if not val:
                continue
            norm = _normalise(val)
            if "opis" in norm or "naziv stavke" in norm:
                return layout

    # 2. B-formula telltale
    sample = rows[:50] if len(rows) >= 50 else rows
    formula_count = 0
    text_count = 0
    for row in sample:
        cell = row[_COL_OPIS] if _COL_OPIS < len(row) else None
        if cell is None or cell.value in (None, ""):
            continue
        if cell.data_type == "f":
            formula_count += 1
        else:
            text_count += 1
    # Ako je >70% B ćelija formula → Layout B
    total = formula_count + text_count
    if total > 5 and formula_count / total > 0.7:
        return "B"

    return "A"


def _parse_arh_sheet(ws, cache: dict) -> list[ParsedItem]:
    """Parsa jedan sheet po _ARH pravilima (Layout A: A=rb, B=opis, C-F=math).

    Layout B (C=rb, D=opis) je legacy predložak — preskačemo ga jer
    nema dovoljno korisnika koji ga još koriste. Marko trenutno radi
    samo Layout A, pa B je outlier."""
    rows = list(ws.iter_rows())
    items: list[ParsedItem] = []
    if not rows:
        return items

    layout = _detect_layout(rows, cache)
    if layout == "B":
        # Layout B je stariji predložak (Westin/Abilia/EmiLu) — skipamo
        return items

    # _ARH Layout A: header u R2 (vidjeli smo: 'red. / broj' | 'opis stavke' | …)
    # Skip prvih ~3 retka — title + header. Heuristika: prvi red gdje
    # A ima rb ili B ima opis bez header riječi.
    header_idx = 0
    for idx, row in enumerate(rows[:10]):
        rb = _str(row, _COL_RB, cache)
        opis = _str(row, _COL_OPIS, cache)
        # Header redovi sadrže "broj"/"opis stavke"/"jed. mjere"/"količina"
        norm = _normalise(opis)
        if norm in ("opis stavke", "opis", "rekapitulacija") or "rekapitul" in norm:
            header_idx = idx + 1
            continue
        if _is_rb(rb):
            header_idx = idx
            break

    blocks: list[_Block] = []
    current = _Block(sheet=ws.title, start_row=header_idx + 1)

    for offset, row in enumerate(rows[header_idx:], start=header_idx + 1):
        rb_str = _str(row, _COL_RB, cache)
        opis_str = _str(row, _COL_OPIS, cache)
        jm = _str(row, _COL_JM, cache)
        kol = _num(_cell_value(row, _COL_KOL, cache))
        cijena = _num(_cell_value(row, _COL_CIJENA, cache))
        iznos_raw = _cell_value(row, _COL_IZNOS, cache)
        iznos = _num(iznos_raw)
        section_marker = _str(row, _COL_J, cache).lower()

        is_empty = not (
            rb_str or opis_str or jm or kol is not None
            or cijena is not None or iznos is not None or section_marker
        )
        if is_empty:
            continue

        # Section marker (sn/st/ss) — pripada current bloku, ne otvara novi
        if section_marker in _SECTION_MARKERS:
            current.section_markers.append({
                "row": offset,
                "kind": section_marker,
                "text": opis_str,
            })
            continue

        # Math row indikatori (jm OR bilo koji od kol/cijena/iznos)
        has_math = bool(jm) or kol is not None or cijena is not None or iznos is not None

        if _is_rb(rb_str):
            # Novi blok — close current, open new
            if not current.is_empty:
                blocks.append(current)
            current = _Block(
                sheet=ws.title,
                rb=rb_str,
                title=opis_str,
                title_row=offset if opis_str else None,
                start_row=offset,
            )
            # Ako rb red ima i math (rijetko, ali moguće), zapiši ga
            if has_math:
                current.math_rows.append({
                    "row": offset,
                    "jm": jm,
                    "kol": kol,
                    "cijena": cijena,
                    "iznos": iznos,
                    "iznos_is_formula": isinstance(iznos_raw, str) and iznos_raw.startswith("="),
                    "position_label": "",
                })
            continue

        # Continuation row inside current block
        if has_math:
            current.math_rows.append({
                "row": offset,
                "jm": jm,
                "kol": kol,
                "cijena": cijena,
                "iznos": iznos,
                "iznos_is_formula": isinstance(iznos_raw, str) and iznos_raw.startswith("="),
                "position_label": opis_str,  # opis na math redu = podstavka label
            })
        elif opis_str:
            current.description_lines.append((offset, opis_str))

    if not current.is_empty:
        blocks.append(current)

    # Emit blocks as ParsedItems
    for block in blocks:
        # Block bez math = section_header (čisti tekst grupa)
        kind = "stavka" if block.math_rows else "section_header"
        text_parts: list[str] = []
        if block.title:
            text_parts.append(block.title)
        text_parts.extend(line for _, line in block.description_lines)
        text = "\n".join(text_parts)
        label = block.title or (text[:120] if text else block.rb)

        items.append(
            ParsedItem(
                position=len(items),
                label=label[:200] if label else "",
                text=text,
                metadata={
                    "sheet": ws.title,
                    "row": block.start_row,
                    "kind": kind,
                    "rb": block.rb,
                    "title": block.title,
                    "title_row": block.title_row,
                    "text_rows": [
                        {"row": r, "text": t} for r, t in block.description_lines
                    ],
                    "math_rows": block.math_rows,
                    "section_markers": block.section_markers,
                    "source": "arh",  # marker da je ovo iz _ARH ground truth
                },
            )
        )

    return items


def parse_arh_xlsx(path: Path) -> ParsedDocument:
    """Parsa _ARHIGON_R0x.xlsx fajl po strogim _ARH pravilima.

    Vraća ParsedDocument sa istim shape-om kao canonical_xlsx parser
    da se rezultati mogu izravno komparirati."""
    try:
        wb = load_workbook(filename=str(path), data_only=False, read_only=False)
        wb_cached = load_workbook(filename=str(path), data_only=True, read_only=False)
    except InvalidFileException as exc:
        raise ParserError(f"Nevažeći _ARH XLSX: {exc}") from exc

    items: list[ParsedItem] = []
    sheets_meta: list[dict[str, Any]] = []

    try:
        for ws in wb.worksheets:
            kind = classify_sheet(ws.title)
            cached_ws = (
                wb_cached[ws.title] if ws.title in wb_cached.sheetnames else None
            )
            cache = _build_formula_cache(cached_ws) if cached_ws is not None else {}

            if kind == "skip":
                sheet_items: list[ParsedItem] = []
            elif kind == "stavke":
                sheet_items = _parse_arh_sheet(ws, cache)
            else:
                # tekst / opci_uvjeti / rekapitulacija — preskačemo,
                # ground truth bitan je samo za stavka sheete.
                sheet_items = []

            for item in sheet_items:
                item.position = len(items)
                items.append(item)

            sheets_meta.append({
                "name": ws.title, "kind": kind, "items": len(sheet_items),
            })
    finally:
        wb.close()
        wb_cached.close()

    return ParsedDocument(
        items=items,
        metadata={"format": "arh-xlsx", "parser": "arh-strict", "sheets": sheets_meta},
    )
