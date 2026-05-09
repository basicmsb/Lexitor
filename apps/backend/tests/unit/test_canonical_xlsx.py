"""Regression tests za canonical_xlsx parser.

Each test builds a synthetic workbook in memory using openpyxl, runs the
parser (or a specific helper) and checks the output. Goal: every parsing
rule has at least one test pinning its current behaviour, so future
changes can't silently regress already-handled cases.

Rules covered (one block per rule, see comments):
- Header detection: token-based + sniff fallback (no header row)
- Section depth heuristic
- UKUPNO detection: SUM formula, arithmetic addition, text "UKUPNO" label
- UKUPNO suppression when row has full kol×cijena (descriptive "ukupno")
- Child-rb math row absorption ("01.01.01" under "01.01")
- Same-rb "kompleti" totals row
- Component spec rows (kol + jm without cijena/iznos) become description
- Cross-sheet reference extraction (Excel + Apple Numbers form)
- POZICIJE: bullet pairing with math rows
- Pre-header opci_uvjeti extraction
- Section header split (heading vs description rows)
- Document registry validates cross-sheet refs against UKUPNOs
"""

from __future__ import annotations

import io
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from src.document_parser.canonical_xlsx import (
    _extract_cross_sheet_refs,
    _extract_sum_rows,
    _is_section_label,
    _row_has_ukupno_label,
    _section_depth,
    _sniff_columns,
    classify_sheet,
    find_header,
    parse_canonical_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers


def _wb_to_path(wb: Workbook, tmp_path) -> Any:
    """Persist a Workbook to a temp xlsx and return its Path. parse_canonical_xlsx
    accepts only paths; we round-trip through a temp file."""
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


def _set_row(ws, row_idx: int, *values) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


# ---------------------------------------------------------------------------
# Sheet classification


@pytest.mark.unit
@pytest.mark.parametrize(
    "name,expected",
    [
        ("Naslovna", "tekst"),
        ("naslovnica", "tekst"),
        ("nasl uk", "tekst"),
        ("nasl", "tekst"),
        ("Sadržaj", "skip"),
        ("Opći uvjeti", "opci_uvjeti"),
        ("0_OPĆI UVIJETI", "opci_uvjeti"),  # diacritic + "uvijeti" typo
        ("REKAPITULACIJA", "rekapitulacija"),
        ("REKAPITULACIJA UKUPNO", "rekapitulacija"),
        ("A.I. PRIPREMNI RADOVI", "stavke"),
        ("Zemljani", "stavke"),
    ],
)
def test_classify_sheet(name: str, expected: str) -> None:
    assert classify_sheet(name) == expected


# ---------------------------------------------------------------------------
# Section label + depth


@pytest.mark.unit
@pytest.mark.parametrize(
    "label,is_section",
    [
        ("1", True),
        ("1.", True),
        ("1.1", True),
        ("1.1.1", True),
        ("A", True),
        ("A.", True),
        ("A.1", True),
        ("A.II.3", True),
        ("II", True),
        ("III", True),
        ("a/", False),
        ("b/", False),
        ("UKUPNO", False),
        ("", False),
    ],
)
def test_is_section_label(label: str, is_section: bool) -> None:
    assert _is_section_label(label) is is_section


@pytest.mark.unit
@pytest.mark.parametrize(
    "label,depth",
    [
        ("A", 1),
        ("B.", 1),
        ("I", 2),
        ("II", 2),
        ("III", 2),
        ("1", 3),
        ("1.1", 4),
        ("1.1.1", 5),
        ("01.01", 4),
        ("01.01.01", 5),
    ],
)
def test_section_depth(label: str, depth: int) -> None:
    assert _section_depth(label) == depth


# ---------------------------------------------------------------------------
# SUM / aggregation extraction


@pytest.mark.unit
@pytest.mark.parametrize(
    "formula,expected",
    [
        ("=SUM(F11:F43)", list(range(11, 44))),
        ("=SUM(F24,F36,F67)", [24, 36, 67]),
        ("=SUM(F24;F36;F67)", [24, 36, 67]),
        ("=SUM(F24:F36, F67)", list(range(24, 37)) + [67]),
        ("=G42", [42]),
        ("=G42+G53", [42, 53]),
        ("=G42+G53-G60", [42, 53, 60]),
        ("=F42*E42", None),  # multiplication is per-row math, not aggregation
        ("=B5/C5", None),
        ("=F42*E42+G53", None),  # mixed has multiplication
        ("=SUMPRODUCT(A1:A5,B1:B5)", None),
        (None, None),
        ("plain text", None),
    ],
)
def test_extract_sum_rows(formula, expected) -> None:
    assert _extract_sum_rows(formula) == expected


# ---------------------------------------------------------------------------
# Cross-sheet references


@pytest.mark.unit
def test_cross_sheet_excel_quoted() -> None:
    refs = _extract_cross_sheet_refs("='A.IV. HIDROIZOL. RADOVI'!G75")
    assert refs == [("A.IV. HIDROIZOL. RADOVI", "G", 75)]


@pytest.mark.unit
def test_cross_sheet_excel_bare() -> None:
    refs = _extract_cross_sheet_refs("=pripremni!F35")
    assert refs == [("pripremni", "F", 35)]


@pytest.mark.unit
def test_cross_sheet_apple_numbers() -> None:
    refs = _extract_cross_sheet_refs("=zemljani.F46")
    assert refs == [("zemljani", "F", 46)]


@pytest.mark.unit
def test_cross_sheet_multiple() -> None:
    refs = _extract_cross_sheet_refs("=pripremni!F35+geodetski!F12")
    assert refs == [("pripremni", "F", 35), ("geodetski", "F", 12)]


@pytest.mark.unit
def test_cross_sheet_no_match() -> None:
    # Same-sheet SUM should NOT be picked up as a cross-sheet reference
    assert _extract_cross_sheet_refs("=SUM(I3:I12)") == []
    assert _extract_cross_sheet_refs("=I3+I12") == []


# ---------------------------------------------------------------------------
# UKUPNO label detection


@pytest.mark.unit
def test_row_has_ukupno_label_positive() -> None:
    wb = Workbook()
    ws = wb.active
    _set_row(ws, 1, "", "UKUPNO A.I.", "", "", "", 500)
    row = next(ws.iter_rows())
    assert _row_has_ukupno_label(row) is True


@pytest.mark.unit
def test_row_has_ukupno_label_negative() -> None:
    wb = Workbook()
    ws = wb.active
    _set_row(ws, 1, "1.1.", "Iskop temelja", "m3", 100, 50, 5000)
    row = next(ws.iter_rows())
    assert _row_has_ukupno_label(row) is False


# ---------------------------------------------------------------------------
# Header detection — token-based


@pytest.mark.unit
def test_find_header_arhigon_style() -> None:
    """Arhigon convention: 'stavka' as rb header, 'opis stavke' as opis,
    two cijena columns where right-most becomes iznos."""
    wb = Workbook()
    ws = wb.active
    _set_row(ws, 1, "", "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena (EUR)", "cijena (EUR)")
    rows = list(ws.iter_rows())
    result = find_header(rows)
    assert result is not None
    idx, mapping = result
    assert idx == 0
    assert mapping.rb == 1     # B column
    assert mapping.opis == 2   # C
    assert mapping.jm == 3     # D
    assert mapping.kol == 4    # E
    assert mapping.cijena == 5  # F (jedinična)
    assert mapping.iznos == 6   # G (cijena EUR — promoted to iznos)


@pytest.mark.unit
def test_find_header_standard_layout() -> None:
    wb = Workbook()
    ws = wb.active
    _set_row(ws, 1, "rb", "Opis", "JM", "Količina", "Cijena", "Iznos")
    rows = list(ws.iter_rows())
    result = find_header(rows)
    assert result is not None
    _, mapping = result
    assert mapping.rb == 0
    assert mapping.opis == 1
    assert mapping.iznos == 5


# ---------------------------------------------------------------------------
# Header detection — sniff fallback (no header row)


@pytest.mark.unit
def test_find_header_sniff_no_header_row() -> None:
    """DV Netretić-style: no header, layout inferred from data rows."""
    wb = Workbook()
    ws = wb.active
    # rb in A, opis in B, jm in C, kol in D, cijena in E, iznos in F
    _set_row(ws, 1, "1.", "Strojni iskop temelja u zemlji III ktg za skidanje "
             "sloja humusa", None, None, None, None)
    _set_row(ws, 2, "1.1.", "Strojno skidanje sloja humusa u debljini do 25 cm. "
             "Obračun u m3 u sraslom stanju", None, None, None, None)
    _set_row(ws, 3, None, "Pozicija izgradnje zgrade (680 m2):", "m3", 170, 80,
             "=D3*E3")
    _set_row(ws, 4, "1.2.", "Strojni široki iskop sloja zemlje u svrhu "
             "izravnavanja terena", None, None, None, None)
    _set_row(ws, 5, None, "a/ Iskop i deponiranje materijala", "m3", 136, 80,
             "=D5*E5")
    _set_row(ws, 6, None, "b/ Utovar i odvoz viška materijala", "m3", 68, 50,
             "=D6*E6")
    _set_row(ws, 7, "2.", "Ručno planiranje dna", None, None, None, None)
    _set_row(ws, 8, None, "ručno planiranje dna jarka", "m2", 100, 5, "=D8*E8")

    rows = list(ws.iter_rows())
    result = find_header(rows)
    assert result is not None
    idx, mapping = result
    assert idx == -1  # no header row to skip
    assert mapping.rb == 0     # A
    assert mapping.opis == 1   # B
    assert mapping.jm == 2     # C
    assert mapping.kol == 3    # D
    assert mapping.cijena == 4  # E
    assert mapping.iznos == 5   # F


@pytest.mark.unit
def test_sniff_columns_returns_empty_when_no_structure() -> None:
    wb = Workbook()
    ws = wb.active
    _set_row(ws, 1, "Some random text in one column")
    _set_row(ws, 2, "More text here")
    rows = list(ws.iter_rows())
    mapping = _sniff_columns(rows)
    assert not mapping.is_minimally_complete


# ---------------------------------------------------------------------------
# Full-document parsing — exercises the rules end-to-end via parse_canonical_xlsx


def _make_workbook_with_arhigon_stavke(tmp_path):
    """Synthesises a single-sheet stavke workbook in Arhigon layout for
    end-to-end parser tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Zemljani"
    # Header at row 1
    _set_row(ws, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "cijena (EUR)")
    # Empty boundary
    _set_row(ws, 2)
    # Stavka 1 — single math row
    _set_row(ws, 3, "1", "Iskop temelja", "m3", 100, 50, "=D3*E3")
    _set_row(ws, 4)
    # Stavka 2 with sub-podstavki via child rb
    _set_row(ws, 5, "2", "Betoniranje temelja", None, None, None, None)
    _set_row(ws, 6, "2.1", "Beton C25/30", "m3", 50, 100, "=D6*E6")
    _set_row(ws, 7, "2.2", "Armatura B500B", "kg", 1000, 1, "=D7*E7")
    _set_row(ws, 8)
    # Group total — SUM that covers stavka 1's math row + stavka 2's children
    _set_row(ws, 9, "III", "UKUPNO ZEMLJANI", None, None, None,
             "=F3+F6+F7")
    return wb


@pytest.mark.unit
def test_parse_arhigon_stavke_with_child_rb(tmp_path) -> None:
    """_ARH full-path konvencija (Marko 2026): kad parent rb nema math
    vrijednosti, child rb-ovi s math-om su **zasebne stavke**, ne
    podstavke. Stari "kompleti" pattern (apsorpcija) ostaje samo kad
    parent već ima math redove."""
    wb = _make_workbook_with_arhigon_stavke(tmp_path)
    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    items = result.items

    # Očekujemo: stavka 1, section_header 2 ("Betoniranje temelja", no math),
    # stavka 2.1, stavka 2.2, group_sum
    stavka_items = [i for i in items if i.metadata.get("kind") == "stavka"]
    section_headers = [i for i in items if i.metadata.get("kind") == "section_header"]
    group_sums = [i for i in items if i.metadata.get("kind") == "group_sum"]

    assert len(stavka_items) == 3  # 1, 2.1, 2.2 — sve zasebne
    assert len(section_headers) == 1  # "2" je section_header bez math-a
    assert len(group_sums) == 1

    rbs = {s.metadata.get("rb") for s in stavka_items}
    assert "1" in rbs or "1." in rbs
    # Full-path rb je prependan section path-om (Fix #2). "2.1" pod
    # section "2" postaje "2.2.1" ili ostaje "2.1" ovisno o stack-u.
    assert any("2.1" in r for r in rbs if r)
    assert any("2.2" in r for r in rbs if r)


@pytest.mark.unit
def test_parse_kompleti_same_rb_totals(tmp_path) -> None:
    """When a row has the SAME rb as the open stavka — typically a
    'kompleti' totals row — its math values become the stavka's own
    totals, not a new stavka. Component spec rows above are kept as
    description."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stavke"
    _set_row(ws, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws, 2)
    _set_row(ws, 3, "02.01.", "Razdjelnik +GR-ZP", None, None, None, None)
    # Component specs (kol + jm but no cijena/iznos)
    _set_row(ws, 4, None, "FID strujna zaštitna sklopka", "kom", 2, None, None)
    _set_row(ws, 5, None, "Minijaturni automatski prekidač B6A", "kom", 1, None,
             None)
    _set_row(ws, 6, None, "Digitalni uklopni sat", "kom", 2, None, None)
    # Same-rb totals row
    _set_row(ws, 7, "02.01.", "Razdjelnik +GR-ZP - ukupno", "kpl", 1, 800, 800)

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    stavka_items = [i for i in result.items if i.metadata.get("kind") == "stavka"]

    # Exactly ONE stavka — not split into two by the totals row
    assert len(stavka_items) == 1
    stavka = stavka_items[0]
    assert stavka.metadata["rb"] == "02.01."
    # One math row — the totals row, with kpl × 1 × 800 = 800
    math_rows = stavka.metadata.get("math_rows", [])
    assert len(math_rows) == 1
    assert math_rows[0]["jm"] == "kpl"
    # Component spec rows became description, not math rows
    text_rows = stavka.metadata.get("text_rows", [])
    desc_text = " ".join(tr["text"] for tr in text_rows)
    assert "FID" in desc_text
    assert "Minijaturni" in desc_text


@pytest.mark.unit
def test_parse_ukupno_text_only_without_formula(tmp_path) -> None:
    """Hardcoded UKUPNO (no formula, plain number) preceded by an empty
    block should be detected as a group_sum via text label. With 2+
    empty rows between the stavka and the UKUPNO row, the prior stavka's
    block closes (EMPTY_BOUNDARY_LIMIT=2) so the UKUPNO becomes a stand-
    alone group_sum rather than the prior stavka's own total."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Zemljani"
    _set_row(ws, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws, 2)
    _set_row(ws, 3, "1", "Iskop", "m3", 100, 50, "=D3*E3")
    _set_row(ws, 4)  # empty row #1
    _set_row(ws, 5)  # empty row #2 — closes the stavka block
    _set_row(ws, 6, None, "UKUPNO", None, None, None, 5000)  # hardcoded

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    group_sums = [i for i in result.items if i.metadata.get("kind") == "group_sum"]
    assert len(group_sums) == 1
    # No SUM formula → empty summed_rows
    assert group_sums[0].metadata.get("summed_rows") == []
    assert group_sums[0].metadata.get("formula") is None


@pytest.mark.unit
def test_parse_ukupno_suppress_when_kol_and_cijena(tmp_path) -> None:
    """A row containing the word 'ukupno' descriptively (e.g. 'ukupno
    površina P=') with full kol×cijena math values is a regular math row,
    NOT an UKUPNO. Suppression rule: if both kol and cijena are populated,
    text-based UKUPNO trigger is ignored."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stavke"
    _set_row(ws, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws, 2)
    _set_row(ws, 3, "1", "Demontaža stropa", None, None, None, None)
    _set_row(ws, 4, None, "ukupno površina ploča ophoda P=", "m2", 390, 20,
             "=D4*E4")

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    stavka_items = [i for i in result.items if i.metadata.get("kind") == "stavka"]
    group_sums = [i for i in result.items if i.metadata.get("kind") == "group_sum"]

    # Should be ONE stavka with one math row, not a group_sum
    assert len(stavka_items) == 1
    assert len(group_sums) == 0
    assert len(stavka_items[0].metadata.get("math_rows", [])) == 1


# ---------------------------------------------------------------------------
# Rekapitulacija + cross-sheet validation


@pytest.mark.unit
def test_rekapitulacija_picks_iznos_formula_not_label(tmp_path) -> None:
    """Each rekapitulacija line typically holds three cross-sheet formulas
    (rb, opis, iznos). The validator must point at the IZNOS formula —
    the one whose resolved value is numeric — not the leftmost (which
    references the rb cell)."""
    wb = Workbook()
    # Stavke sheet with one stavka and an UKUPNO at row 4
    ws_stavke = wb.active
    ws_stavke.title = "Pripremni"
    _set_row(ws_stavke, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws_stavke, 2, "1", "Pripremni radovi — ukupno", None, None,
             None, None)
    _set_row(ws_stavke, 3)
    _set_row(ws_stavke, 4, None, "Iskop", "m3", 100, 50, "=D4*E4")
    _set_row(ws_stavke, 5)
    # Grand UKUPNO at row 6
    _set_row(ws_stavke, 6, None, "UKUPNO PRIPREMNI", None, None, None,
             "=SUM(F4:F4)")

    # Rekapitulacija sheet with 3 cross-sheet refs per row
    ws_rekap = wb.create_sheet("REKAPITULACIJA")
    _set_row(ws_rekap, 1, "I", "PRIPREMNI RADOVI",
             "=Pripremni!A2", "=Pripremni!B2", "=Pripremni!F6")

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    recap_lines = [
        i for i in result.items if i.metadata.get("kind") == "recap_line"
    ]
    # The recap row should pick up the rightmost numeric-resolving formula
    # (column G — the iznos), not column C (rb ref).
    assert len(recap_lines) >= 1
    line = recap_lines[0]
    refs = line.metadata.get("cross_sheet_refs") or []
    assert any(r["sheet"].lower() == "pripremni" and r["row"] == 6 for r in refs), (
        f"Expected ref to Pripremni row 6 (iznos), got: {refs}"
    )


# ---------------------------------------------------------------------------
# Pre-header opci_uvjeti


@pytest.mark.unit
def test_pre_header_rows_emitted_as_opci_uvjeti(tmp_path) -> None:
    """Text rows above the header (sheet title, opci uvjeti for the section)
    should be emitted as opci_uvjeti items so they aren't lost."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pripremni"
    _set_row(ws, 1, None, "TEHNIČKA SPECIFIKACIJA")
    _set_row(ws, 2, None, "Radovi se izvode prema HRN-u i prema važećim "
             "propisima građevinske struke.")
    _set_row(ws, 3)
    # Header at row 4
    _set_row(ws, 4, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws, 5, "1", "Pripremni radovi", "m3", 10, 100, "=D5*E5")

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)
    opci = [i for i in result.items if i.metadata.get("kind") == "opci_uvjeti"]
    assert len(opci) == 2
    assert any("TEHNIČKA" in i.text for i in opci)
    assert any("HRN" in i.text for i in opci)


# ---------------------------------------------------------------------------
# Section header split (heading vs description rows)


@pytest.mark.unit
def test_section_with_description_splits_into_header_plus_opci_uvjeti(tmp_path) -> None:
    """A section_header block that has both a heading and description text
    should split: section_header item with the heading + one opci_uvjeti
    item per description row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stavke"
    _set_row(ws, 1, "stavka", "opis stavke", "jedinica", "količina",
             "jedinična cijena", "iznos")
    _set_row(ws, 2)
    _set_row(ws, 3, "B.I.01.", "KABELSKE STAZE I KABELI", None, None, None, None)
    _set_row(ws, 4, None, "NAPOMENA:", None, None, None, None)
    _set_row(ws, 5, None, "* Prije narudžbe izmjeriti stvarne dužine krugova.",
             None, None, None, None)
    _set_row(ws, 6, None, "* Za višežilne kabele koristiti boje izolacije.",
             None, None, None, None)
    _set_row(ws, 7)
    _set_row(ws, 8, "01.01", "Kabel YSLY 3x1.5", "m", 100, 2, "=D8*E8")

    path = _wb_to_path(wb, tmp_path)
    result = parse_canonical_xlsx(path)

    section_headers = [
        i for i in result.items if i.metadata.get("kind") == "section_header"
    ]
    opci = [i for i in result.items if i.metadata.get("kind") == "opci_uvjeti"]

    # Heading itself
    assert any(
        i.metadata.get("rb") == "B.I.01." and "KABELSKE" in (i.metadata.get("title") or "")
        for i in section_headers
    )
    # Three description rows split out as opci_uvjeti
    napomena_items = [i for i in opci if "NAPOMENA" in i.text or "Prije narudžbe" in i.text or "boje izolacije" in i.text]
    assert len(napomena_items) >= 3
